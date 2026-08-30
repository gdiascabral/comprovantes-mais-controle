# -*- coding: utf-8 -*-
"""
Anexar Comprovantes — Mais Controle

Fluxo da janela:
  1) Informe o PERÍODO (data de pagamento dos comprovantes) e a PASTA dos
     PDFs renomeados (padrão "VALOR - DESCRIÇÃO - DATA") e clique em
     "Carregar contas" — o app abre o Chrome, ENTRA SOZINHO no Mais Controle
     (com a senha guardada no 🔑 Login) e busca os títulos PAGOS do período.
  2) Marque as CONTAS BANCÁRIAS desejadas (caixas de seleção).
  3) "Casar e anexar": verifica quem já tem comprovante (pula), casa os
     pendentes com os PDFs e anexa. No fim, gera um relatório Excel.

O botão "Abrir o Mais Controle" não é mais um passo: serve para o primeiro
acesso (quando ainda não há senha guardada) e para destravar sessão caída.

Modo alternativo "Por lista": anexa a partir de um CSV (launchId,valor,arquivo_pdf)
ou de um Excel com aba CERTEZA (coluna link + PDF(s)).
"""
import os
import queue
import time
import traceback
from concurrent.futures import ThreadPoolExecutor
from threading import Event
from datetime import date, datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from . import config, matcher, mc_api, planilha, credenciais
    from .mc_client import MCClient, SemRede
except ImportError:
    import config, matcher, mc_api, planilha, credenciais
    from mc_client import MCClient, SemRede

try:                                     # utilitários compartilhados (raiz)
    import util
except ModuleNotFoundError:              # rodando este módulo isoladamente
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    import util

LINK = config.MC_URL_LANCAMENTO
_fmt_dur = util.fmt_dur
_norm = util.norm


#: Formatos partilhados com a Conferência. Moram no util.py para a
#: Conferência não precisar importar nome PRIVADO daqui — import assim cria
#: dependência de mão única entre duas abas e quebra quando uma é reordenada.
_data_api = util.data_api
_fmt_val = util.fmt_val


def _texto_do_erro(e: Exception) -> str:
    """UMA frase para o Registro; o traceback vai para o `diagnostico.log`.

    Falta de internet é recado, não defeito do app: mostra só a orientação.

    O resto despejava o `traceback.format_exc()` inteiro dentro do Registro —
    vinte linhas de Python no campo que a pessoa deveria ler para saber o que
    fazer, e a frase que importava ficava no meio delas. O traceback continua
    existindo, no lugar que o `config.diag` foi feito para guardar: engolir o
    erro na tela, mas deixar o motivo gravado. Só chame de dentro de um
    `except` — é de lá que o `format_exc()` tira o que registrar."""
    if isinstance(e, SemRede):
        return "⚠ " + str(e)
    config.diag(f"{e!r}\n{traceback.format_exc()}")
    return "ERRO: " + (str(e) or
                       "falha sem mensagem — o motivo está no diagnostico.log")


def _resumo_cands(pe: dict) -> str:
    """Candidatos que sobraram para um pagamento em dúvida, do mais provável
    para o menos, com o que bateu em cada um — mesmo detalhe que a janela."""
    partes = []
    for c in sorted((c for c in pe["cands"] if c["pdf"]["used_by"] is None),
                    key=lambda c: -c["score"]):
        sinais = " + ".join(s for s in ("OC/NF" if c["ocnf"] else "",
                                        "centro de custo" if c["cc"] else "",
                                        "data" if c["date"] else "") if s)
        partes.append(f"{c['pdf']['fn']}  [{sinais or 'só o valor'}]")
    return " || ".join(partes) or "(sem candidatos livres)"


def _abrir_url(url: str):
    """Abre uma URL no navegador padrão. Usa os.startfile (sempre presente no
    executável); só recorre ao módulo webbrowser se aquele falhar — assim o
    app não depende do webbrowser estar embutido no motor."""
    try:
        os.startfile(url)                # Windows: abre no navegador padrão
    except Exception:
        try:
            import webbrowser
            webbrowser.open(url)
        except Exception:
            pass


try:                                     # widgets compartilhados (raiz)
    import widgets
except ModuleNotFoundError:              # rodando este módulo isoladamente
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    import widgets

#: O campo de data mora em widgets.py e é usado por TODAS as abas que pedem
#: data. Ficava aqui dentro, e a Conferência tinha de importá-lo desta aba —
#: uma aba dependendo de outra só para reaproveitar um Entry.
CampoData = widgets.CampoData


class AnexarFrame(ttk.Frame):
    """Conteúdo do app Anexar Comprovantes (usável sozinho ou como aba)."""

    def __init__(self, master):
        super().__init__(master)
        self.q = queue.Queue()
        self.worker = None
        # TODO o trabalho com o navegador roda nesta ÚNICA thread (exigência
        # do Playwright: os objetos só podem ser usados na thread que os criou).
        self.exec = ThreadPoolExecutor(max_workers=1)
        self._pausa = Event()   # ⏸ pausado enquanto setado
        self._parar = Event()   # ⏹ interrompe o processo em andamento
        # Events em que a thread do navegador está BLOQUEADA esperando gente
        # (hoje só a janela de dúvidas). `fechar()` precisa soltá-los, senão a
        # thread nunca termina e o processo não morre ao fechar a janela.
        self._esperas: set = set()
        # Quem está com o navegador agora (ver `submeter`/`ocupado`).
        self._trabalho_atual = None
        self._rotulo_atual = None
        self.mc = None                       # MCClient aberto entre as etapas
        self.api = None
        self.ultimo_relatorio = None
        self.pagos = []                      # registros de montar_pagos()
        self.vars_contas: dict[str, tk.BooleanVar] = {}

        hoje = date.today()
        self.v_ini = tk.StringVar(value=hoje.replace(day=1).strftime("%d/%m/%Y"))
        self.v_fim = tk.StringVar(value=hoje.strftime("%d/%m/%Y"))
        self.v_pasta = tk.StringVar()
        self.v_lista = tk.StringVar()
        self.v_modo = tk.StringVar(value="auto")
        self.v_dry = tk.BooleanVar(value=False)
        self.v_ign = tk.BooleanVar(value=True)
        self.v_ign_ap = tk.BooleanVar(value=True)
        self._build()
        try:                             # já nasce na cor do tema (sem flash)
            self.aplicar_cores(util.cor_escura(ttk.Style().lookup("TFrame", "background")))
        except Exception:
            pass
        self.after(150, self._drain)

    # ---------------------------------------------------------------- layout
    def _build(self):
        PADX = widgets.PADX

        # ---- cabeçalho
        self.cab = widgets.Cabecalho(
            self, "Anexar Comprovantes",
            "Busca os pagos do período, descobre quem está sem comprovante "
            "e anexa o PDF certo em cada um.",
            trilha="Comprovantes  ›  Anexar")
        self.cab.pack(fill="x", padx=PADX, pady=(16, 12))

        # Os dois passos do fluxo vão para o cabeçalho, e o segundo é o verde:
        # anexar é o que esta tela existe para fazer. O resto (login, pausar,
        # parar, abrir relatório) fica embaixo, junto da barra de execução —
        # não são passos, e ficavam do mesmo tamanho dos que são.
        self.b1 = widgets.Botao(self.cab.acoes, "Carregar contas",
                                papel="passo", command=self.conectar)
        self.b1.pack(side="left", padx=(0, 8))
        self.b2 = widgets.Botao(self.cab.acoes, "Casar e anexar", papel="acao",
                                command=self.executar, state="disabled")
        self.b2.pack(side="left")

        # Os cartões passam a ser numerados, e os botões deixam de ser: era o
        # "▶ 1." no botão e a trilha de passos contando a mesma coisa duas
        # vezes. Agora o número está num lugar só.
        self.f_auto = widgets.Cartao(self, "Período e pasta dos comprovantes", 1)
        self.f_auto.pack(fill="x", padx=PADX, pady=(0, 12))
        fa = self.f_auto
        linha = ttk.Frame(fa)
        linha.pack(fill="x")
        widgets.Campo(linha, "Pagamento de",
                      lambda p: CampoData(p, self.v_ini)
                      ).pack(side="left", padx=(0, 16))
        widgets.Campo(linha, "Até", lambda p: CampoData(p, self.v_fim)
                      ).pack(side="left")

        pasta = ttk.Frame(fa)
        pasta.pack(fill="x", pady=(12, 0))
        ttk.Label(pasta, text="PASTA DOS PDFs RENOMEADOS", style="Rotulo.TLabel"
                  ).pack(anchor="w", pady=(0, 3))
        campo = ttk.Frame(pasta)
        campo.pack(fill="x")
        ttk.Entry(campo, textvariable=self.v_pasta).pack(side="left", fill="x",
                                                         expand=True)
        widgets.Botao(campo, "Selecionar…", papel="neutro",
                      command=lambda: self.v_pasta.set(
                          (filedialog.askdirectory() or self.v_pasta.get())
                          .replace("\\", "/"))).pack(side="left", padx=(8, 0))

        # As caixas de marcar ficam agrupadas EMBAIXO do formulário, e não
        # espalhadas entre os campos: elas são exceções à regra do lote, e
        # lidas em bloco dá para conferir as duas de uma vez.
        marcas = ttk.Frame(fa)
        marcas.pack(fill="x", pady=(12, 0))
        ttk.Checkbutton(marcas, variable=self.v_ign,
                        text="Ignorar tarifas bancárias, IOF, cesta e pacote "
                             "de serviços").pack(anchor="w")
        ttk.Checkbutton(marcas, variable=self.v_ign_ap,
                        text="Ignorar aportes de capital e distribuição de "
                             "lucros").pack(anchor="w", pady=(4, 0))

        # ---- escolha do modo (entre os blocos 1 e 2)
        self.topo = widgets.Cartao(self, "Como casar comprovante e lançamento", 2)
        self.topo.pack(fill="x", padx=PADX, pady=(0, 12))
        ttk.Radiobutton(self.topo, text="Automático — casar pelos nomes dos PDFs",
                        variable=self.v_modo, value="auto",
                        command=self._alternar_modo).pack(anchor="w")
        ttk.Radiobutton(self.topo, text="Por lista pronta (.csv / .xlsx)",
                        variable=self.v_modo, value="lista",
                        command=self._alternar_modo).pack(anchor="w", pady=(4, 0))

        # ---- card: contas
        self.f_contas = widgets.Cartao(self, "Contas bancárias — marque as desejadas", 3)
        self.f_contas.pack(fill="x", padx=PADX, pady=(0, 12))
        self.rodape_contas = widgets.RodapeTabela(self.f_contas.acoes)
        self.rodape_contas.pack()
        self.contas_box = ttk.Frame(self.f_contas)
        self.contas_box.pack(fill="x")
        ttk.Label(self.contas_box, style="Tenue.TLabel",
                  text='Clique em "Carregar contas" para listar as contas.'
                  ).pack(anchor="w")

        # ---- card: modo lista (mostrado só no modo "Por lista")
        self.f_lista = widgets.Cartao(self, "Lista pronta")
        fl = self.f_lista
        ttk.Label(fl, text="ARQUIVO (.csv ou .xlsx)", style="Rotulo.TLabel"
                  ).pack(anchor="w", pady=(0, 3))
        escolha = ttk.Frame(fl)
        escolha.pack(fill="x")
        ttk.Entry(escolha, textvariable=self.v_lista).pack(side="left", fill="x",
                                                           expand=True)
        widgets.Botao(escolha, "Selecionar…", papel="neutro",
                      command=self._sel_lista).pack(side="left", padx=(8, 0))

        # ---- barra de execução e o que não é passo
        acao = ttk.Frame(self, style="Fundo.TFrame")
        acao.pack(fill="x", padx=PADX, pady=(0, 10))
        btns = ttk.Frame(acao, style="Fundo.TFrame")
        btns.pack(side="right", padx=(16, 0))
        ttk.Checkbutton(btns, text="Simular", variable=self.v_dry,
                        style="Fundo.TCheckbutton").pack(side="left", padx=(0, 10))
        self.b_pause = widgets.Botao(btns, "⏸  Pausar", papel="neutro",
                                     command=self._pausar_toggle,
                                     state="disabled")
        self.b_pause.pack(side="left")
        self.b_stop = widgets.Botao(btns, "⏹  Parar", papel="perigo",
                                    command=self._parar_click, state="disabled")
        self.b_stop.pack(side="left", padx=(8, 0))
        self.b_rel = widgets.Botao(btns, "📄  Abrir relatório", papel="neutro",
                                   command=self._abrir_relatorio,
                                   state="disabled")
        self.b_rel.pack(side="left", padx=(8, 0))
        # "Abrir e acessar" deixou de ser passo: com o login salvo o app entra
        # sozinho. O botão continua aqui, sem número, para o primeiro acesso
        # (quando ainda não há senha guardada) e para destravar sessão caída.
        self.b0 = widgets.Botao(btns, "Abrir o Mais Controle", papel="neutro",
                                command=self.abrir_mc)
        self.b0.pack(side="left", padx=(8, 0))
        self.b_login = widgets.Botao(btns, "🔑  Login", papel="neutro",
                                     command=self._gerenciar_login)
        self.b_login.pack(side="left", padx=(8, 0))

        self.barra_exec = widgets.BarraExecucao(acao)
        self.barra_exec.pack(side="left", fill="x", expand=True)
        # `lbl` e `pb` continuam existindo com os nomes de sempre: o `_drain` e
        # as chamadas de progresso não sabem que a barra virou outro widget.
        self.lbl = self.barra_exec.lbl
        self.pb = self.barra_exec.pb

        # ---- card: registro (cresce quando tem o que mostrar)
        self.reg = widgets.Cartao(self, "Registro", padding=(12, 10))
        self.reg.pack(fill="x", padx=PADX, pady=(0, 12))
        self.log = tk.Text(self.reg, wrap="word", relief="flat", borderwidth=0,
                           highlightthickness=0)
        self.log.pack(fill="both", expand=True)
        widgets.estilo_log(self.log)
        self._mostrar_placeholder()
        widgets.registro_elastico(self.reg, self.log)
        self._alternar_modo()

    def _mostrar_placeholder(self):
        self.log.delete("1.0", "end")
        self.log.insert("end", "\n\n", "ph")
        self.log.insert("end", "O andamento e os resultados aparecerão aqui.\n", "ph")
        self.log.insert("end", "\nSiga os dois passos no alto da tela — o app "
                               "abre o Mais Controle e entra sozinho.\n", "ph")
        self._ph = True

    def _alternar_modo(self):
        if self.v_modo.get() == "auto":
            self.f_lista.pack_forget()
            self.b1.config(state="normal")
        else:
            self.f_lista.pack(fill="x", padx=widgets.PADX, pady=(0, 12),
                              after=self.topo)
            self.b2.config(state="normal")

    def _sel_lista(self):
        f = filedialog.askopenfilename(filetypes=[("Lista", "*.csv *.xlsx")])
        if f:
            self.v_lista.set(f.replace("\\", "/"))

    def _log(self, msg):
        self.q.put(("log", msg))

    # ---------------------------------------------------------------- etapa 1
    def abrir_mc(self):
        if self.worker and not self.worker.done():
            return
        self.b0.config(state="disabled")
        self.worker = self.submeter("Abrir o Mais Controle", self._t_abrir)

    # ------------------------------------------- navegador compartilhado
    def submeter(self, rotulo: str, fn, *a, dona=None, **k):
        """Manda trabalho para a thread do navegador, registrando o dono.

        Seis abas dividem UM navegador e UMA thread (o Playwright síncrono não
        aceita mais, e o ERP só admite uma sessão por usuário). Sem registrar
        quem está usando, clicar numa segunda aba enquanto a primeira trabalha
        apenas ENFILEIRAVA o pedido: a tela não dizia nada e a segunda tarefa
        começava sozinha vários minutos depois, quando a pessoa já tinha
        desistido e ido fazer outra coisa.

        `dona` é a ABA que pediu — o `rotulo` só descreve a tarefa. A barra
        lateral precisa das duas coisas: qual item marcar com ● e o que
        escrever no rodapé. Sem `dona`, o único jeito de descobrir a aba seria
        adivinhar pelo texto do rótulo."""
        fut = self.exec.submit(fn, *a, **k)
        self._trabalho_atual = fut
        self._rotulo_atual = rotulo
        self._dona_atual = dona if dona is not None else self
        return fut

    def ocupado(self) -> str | None:
        """Rótulo da tarefa que está com o navegador agora, ou None."""
        fut = self._trabalho_atual
        if fut is not None and not fut.done():
            return self._rotulo_atual or "outra tarefa"
        return None

    def dona_ocupada(self):
        """A ABA que está com o navegador agora, ou None."""
        fut = self._trabalho_atual
        if fut is not None and not fut.done():
            return getattr(self, "_dona_atual", self)
        return None

    def avisar_se_ocupado(self, dona: str) -> bool:
        """True (e mostra o aviso) se o navegador já estiver em uso."""
        quem = self.ocupado()
        if not quem:
            return False
        messagebox.showinfo(
            "Navegador ocupado",
            f"O navegador está ocupado com: {quem}.\n\n"
            f"O Mais Controle aceita uma sessão por usuário, então "
            f"{dona} precisa esperar terminar.")
        return True

    def garantir_sessao(self, log=None):
        """Abre o Chrome e prepara a API, se ainda não estiverem prontos.
        Deve rodar na thread self.exec (a dona do navegador).

        Duas coisas que faltavam e davam erro longe da causa:

        1. o retorno de `garantir_login()` era ignorado. Sem sessão, o ERP
           responde com a grade VAZIA em vez de erro, e o que chegava ao
           usuário era "a grade não carregou nenhuma linha" — que parece
           layout mudado e é sessão faltando;
        2. com `self.mc` já aberto, ninguém revalidava nada. Mas o ERP aceita
           UMA sessão por usuário: a API de saldos da Conciliação derruba a do
           navegador, e a aba seguinte reaproveitava um Chrome deslogado.
        """
        log = log or self._log

        # NAVEGADOR MORTO É OUTRA COISA DE SESSÃO CAÍDA, e tratá-los igual era
        # o que obrigava a fechar o app para trocar de aba. Quem fecha a janela
        # do Chrome no X — ou tem o Chrome morto por outro motivo — deixa este
        # objeto apontando para nada: a aba seguinte chamava `esta_logado()`
        # nele, tomava erro de Playwright, e não havia caminho de volta sem
        # reiniciar. Aqui o app simplesmente abre outro.
        if self.mc is not None and not self.mc.vivo():
            log("A janela do Chrome não está mais aberta — abrindo outra...")
            try:
                self.mc.__exit__(None, None, None)
            except Exception:
                pass                      # já estava morto; o que importa é soltar
            self.mc = None
            self.api = None

        if self.mc is None:
            log("Abrindo o Chrome e entrando no Mais Controle...")
            self.mc = MCClient(log=log).__enter__()
            self.api = mc_api.MCApi(self.mc)
            if not self.mc.garantir_login():
                raise RuntimeError(
                    "não consegui entrar no Mais Controle.\n"
                    "Sem a sessão do navegador o ERP devolve as telas vazias, "
                    "e o resultado sairia errado — por isso parei aqui.\n"
                    "Entre na janela do Chrome que abriu e rode de novo.")
        elif not self.mc.esta_logado():
            log("A sessão do Mais Controle caiu — entrando de novo...")
            if not self.mc.garantir_login():
                raise RuntimeError(
                    "a sessão do Mais Controle caiu e não consegui refazer.\n"
                    "Entre na janela do Chrome que está aberta e rode de novo.")
        return self.api

    def _gerenciar_login(self):
        """Botão 🔑 Login: cadastrar/trocar/remover o login salvo."""
        def done(creds):
            if creds and creds[2]:
                try:
                    credenciais.salvar(creds[0], creds[1])
                    self.lbl.config(text="Login salvo neste computador.")
                except Exception as e:
                    messagebox.showerror("Erro", f"Não consegui salvar o login:\n{e}")
        self._mostrar_dialogo_login(done)

    def _mostrar_dialogo_login(self, on_done):
        """Diálogo de login (thread principal). Chama on_done((email, senha,
        salvar)) ao confirmar, ou on_done(None) ao cancelar."""
        top = tk.Toplevel(self)
        top.title("Login do Mais Controle")
        top.transient(self.winfo_toplevel())
        top.resizable(False, False)
        widgets.barra_de_titulo(top)
        top.attributes("-topmost", True)     # fica à frente da janela do Chrome
        top.lift()
        top.after(50, top.focus_force)
        salvos = credenciais.carregar()
        ttk.Label(top, wraplength=380, justify="left",
                  text="Entre com seu login do Mais Controle. Ele é salvo cifrado "
                       "neste computador (login automático nas próximas vezes)."
                  ).grid(row=0, column=0, columnspan=2, padx=14, pady=(14, 10), sticky="w")
        v_email = tk.StringVar(value=(salvos[0] if salvos else ""))
        v_senha = tk.StringVar(value=(salvos[1] if salvos else ""))
        v_salvar = tk.BooleanVar(value=True)
        ttk.Label(top, text="E-mail").grid(row=1, column=0, sticky="w", padx=14)
        ttk.Entry(top, textvariable=v_email, width=42
                  ).grid(row=2, column=0, columnspan=2, padx=14, sticky="we")
        ttk.Label(top, text="Senha").grid(row=3, column=0, sticky="w", padx=14, pady=(8, 0))
        e_senha = ttk.Entry(top, textvariable=v_senha, width=42, show="•")
        e_senha.grid(row=4, column=0, columnspan=2, padx=14, sticky="we")
        ttk.Checkbutton(top, text="Salvar neste computador (entrar automaticamente)",
                        variable=v_salvar).grid(row=5, column=0, columnspan=2,
                                                sticky="w", padx=14, pady=10)
        bar = ttk.Frame(top)
        bar.grid(row=6, column=0, columnspan=2, sticky="we", padx=14, pady=(0, 14))

        def concluir(creds):
            top.destroy()
            on_done(creds)

        def confirmar():
            email, senha = v_email.get().strip(), v_senha.get()
            if not email or not senha:
                messagebox.showwarning("Login", "Preencha e-mail e senha.")
                return
            concluir((email, senha, v_salvar.get()))

        b_ok = widgets.Botao(bar, "Entrar", papel="acao", command=confirmar)
        b_ok.pack(side="right")
        widgets.Botao(bar, "Cancelar", papel="neutro",
                      command=lambda: concluir(None)
                      ).pack(side="right", padx=(0, 8))
        if salvos:
            ttk.Button(bar, text="Remover login salvo",
                       command=lambda: (credenciais.apagar(), concluir(None))
                       ).pack(side="left")
        top.protocol("WM_DELETE_WINDOW", lambda: concluir(None))
        e_senha.focus_set()
        top.grab_set()

    def _t_abrir(self):
        inicio = time.time()
        self._log(f"⏱ Etapa 1 — início: {time.strftime('%H:%M:%S')}")
        try:
            self.garantir_sessao()
            self._log("Mais Controle aberto. Agora confira o período e a pasta dos "
                      "PDFs e clique em \"1. Carregar contas\".")
            self._log(f"⏱ Etapa 1 — fim: {time.strftime('%H:%M:%S')} "
                      f"({_fmt_dur(time.time() - inicio)})")
        except Exception as e:
            self._log(_texto_do_erro(e))
            # Fecha o Chrome ANTES de soltar a referência. Sem isto o processo
            # ficava órfão segurando o perfil, e a tentativa seguinte batia em
            # "profile is already in use" — erro que não lembra em nada a causa.
            try:
                if self.mc:
                    self.mc.__exit__(None, None, None)
            except Exception as e2:
                config.diag(f"_t_abrir: o Chrome não fechou depois do erro: {e2!r}")
            self.mc = None
            self.api = None
        self.q.put(("reabilitar0", None))

    def aplicar_cores(self, escuro: bool):
        """Cor do registro. As legendas seguem o tema sozinhas.

        Só o `tk.Text` precisa disso: ele é widget clássico e não tem estilo
        do ttk. Tudo que virou "Apoio.TLabel" se reconfigura em `widgets`."""
        try:
            widgets.estilo_log(self.log, escuro)
        except tk.TclError:
            pass

    # -------------------------------------------------------- pausar / parar
    def _pausar_toggle(self):
        if self._pausa.is_set():
            self._pausa.clear()
            self.b_pause.config(text="⏸ Pausar")
            self.lbl.config(text="Retomando...")
        else:
            self._pausa.set()
            self.b_pause.config(text="▶ Continuar")
            self.lbl.config(text="⏸ Pausado.")

    def _parar_click(self):
        self._parar.set()
        self._pausa.clear()
        self.b_pause.config(text="⏸ Pausar")
        self.lbl.config(text="Parando...")

    def _checar_pausa(self) -> bool:
        """Bloqueia enquanto pausado; retorna True se o usuário mandou parar."""
        while self._pausa.is_set() and not self._parar.is_set():
            time.sleep(0.3)
        return self._parar.is_set()

    # ------------------------------------------------------------ relatório
    def _abrir_relatorio(self):
        if self.ultimo_relatorio and Path(self.ultimo_relatorio).exists():
            try:
                os.startfile(self.ultimo_relatorio)
            except OSError as e:
                messagebox.showerror("Erro", f"Não consegui abrir o relatório:\n{e}")
        else:
            messagebox.showinfo("Relatório", "Nenhum relatório gerado ainda.")

    # ------------------------------------------------------ resolver dúvidas
    def _janela_duvidas(self, duvidas, ev):
        """Janela para o usuário escolher o PDF certo dos casamentos em dúvida.

        Mostra o pagamento INTEIRO (descrição sem cortar, centro de custo, nº
        doc, categoria, conta) e, para cada PDF candidato, o que bateu e o que
        não bateu — é isso que permite decidir sem abrir o Mais Controle. Dá
        também para abrir o PDF na hora, para conferir o comprovante."""
        pasta = Path(self.v_pasta.get() or ".")
        top = tk.Toplevel(self)
        top.title(f"Resolver dúvidas ({len(duvidas)})")
        top.transient(self.winfo_toplevel())
        widgets.barra_de_titulo(top)
        try:
            top.geometry(f"{min(1180, self.winfo_screenwidth() - 80)}"
                         f"x{min(780, self.winfo_screenheight() - 120)}")
        except tk.TclError:
            pass
        ttk.Label(top, wraplength=1120, justify="left",
                  text=f"O app não teve certeza em {len(duvidas)} pagamento(s): "
                       "havia mais de um PDF com o mesmo valor, ou nenhum sinal "
                       "forte (OC/NF, centro de custo, data) para decidir. "
                       "Escolha o PDF certo em cada um — ou deixe em dúvida "
                       "para decidir depois pelo relatório.\n"
                       "Dica: dê dois cliques na linha para abrir o PDF."
                  ).pack(anchor="w", padx=12, pady=(10, 6))

        rodape = ttk.Frame(top)
        rodape.pack(side="bottom", fill="x", padx=12, pady=10)

        canvas = tk.Canvas(top, highlightthickness=0)
        barra = ttk.Scrollbar(top, orient="vertical", command=canvas.yview)
        quadro = ttk.Frame(canvas)
        janela = canvas.create_window((0, 0), window=quadro, anchor="nw")
        quadro.bind("<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfigure(janela, width=e.width))
        canvas.configure(yscrollcommand=barra.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=4)
        barra.pack(side="right", fill="y")
        # Roda do mouse SÓ enquanto o ponteiro está sobre esta janela.
        # `bind_all` sequestrava o evento do app inteiro: com a janela de
        # dúvidas aberta (ou depois dela, se `concluir` não rodasse), rolar
        # qualquer outra aba mexia nesta lista.
        def _roda(e):
            canvas.yview_scroll(-(e.delta // 120), "units")

        canvas.bind("<Enter>", lambda e: top.bind("<MouseWheel>", _roda))
        canvas.bind("<Leave>", lambda e: top.unbind("<MouseWheel>"))

        def _abrir_pdf(tv, mapa):
            pd = mapa.get((tv.selection() or [None])[0])
            if pd is None:
                messagebox.showinfo("Abrir PDF",
                                    "Selecione um dos arquivos da lista.")
                return
            alvo = pasta / pd["fn"]
            try:
                os.startfile(str(alvo))
            except OSError as e:
                messagebox.showerror("Erro", f"Não consegui abrir:\n{alvo}\n\n{e}")

        escolhas = []
        for pe in duvidas:
            cands = sorted((c for c in pe["cands"] if c["pdf"]["used_by"] is None),
                           key=lambda c: -c["score"])
            vals = sorted(set(pe.get("valores") or [pe["valor"]]))
            titulo = f" R$ {_fmt_val(pe['valor'])}"
            outros = [v for v in vals if v != pe["valor"]]
            if outros:
                titulo += " (pago; nominal " + ", ".join(_fmt_val(v) for v in outros) + ")"
            titulo += f" — {pe['dataFull']} — {pe['conta']} "
            bloco = ttk.LabelFrame(quadro, text=titulo)
            bloco.pack(fill="x", padx=4, pady=6)

            doc = pe["doc"] or "—"
            if pe.get("ocs"):
                doc += "     OC/NF: " + ", ".join(pe["ocs"])
            for rotulo, texto in (("Favorecido", pe.get("favorecido") or "—"),
                                  ("Descrição", pe["desc"] or "(sem descrição)"),
                                  ("Centro de custo", "; ".join(pe["works"]) or "—"),
                                  ("Nº doc", doc),
                                  ("Categoria", pe.get("categoria") or "—")):
                ttk.Label(bloco, text=f"{rotulo}: {texto}", wraplength=1080,
                          justify="left").pack(anchor="w", padx=8)

            ttk.Label(bloco, text=f"{len(cands)} PDF(s) livre(s) com esse valor:"
                      ).pack(anchor="w", padx=8, pady=(8, 2))
            tv = ttk.Treeview(bloco, columns=("sinais", "arquivo", "data", "desc"),
                              show="headings", selectmode="browse",
                              height=min(max(len(cands) + 1, 2), 7))
            for col, cab, larg, estica in (("sinais", "O que bateu", 190, False),
                                           ("arquivo", "Arquivo", 470, False),
                                           ("data", "Data do PDF", 85, False),
                                           ("desc", "Descrição do PDF", 290, True)):
                tv.heading(col, text=cab)
                tv.column(col, width=larg, anchor="w", stretch=estica)
            tv.insert("", "end", iid="_nada",
                      values=("—", "(deixar em dúvida)", "", ""))
            mapa = {}
            for k, c in enumerate(cands):
                pd = c["pdf"]
                sinais = " ".join(s for s in ("✔ OC/NF" if c["ocnf"] else "",
                                              "✔ centro de custo" if c["cc"] else "",
                                              "✔ data" if c["date"] else "") if s)
                dt = pd["data"]
                tv.insert("", "end", iid=f"c{k}",
                          values=(sinais or "só o valor bate", pd["fn"],
                                  f"{dt[:2]}/{dt[2:]}" if dt else "—", pd["desc"]))
                mapa[f"c{k}"] = pd
            tv.selection_set("_nada")
            tv.pack(fill="x", padx=8, pady=(0, 4))
            tv.bind("<Double-1>", lambda e, t=tv, m=mapa: _abrir_pdf(t, m))

            botoes = ttk.Frame(bloco)
            botoes.pack(fill="x", padx=8, pady=(0, 6))
            ttk.Button(botoes, text="Abrir lançamento no navegador",
                       command=lambda i=pe["launchId"]: _abrir_url(LINK + str(i))
                       ).pack(side="right")
            ttk.Button(botoes, text="📄  Abrir PDF selecionado",
                       command=lambda t=tv, m=mapa: _abrir_pdf(t, m)
                       ).pack(side="right", padx=(0, 8))
            escolhas.append((pe, tv, mapa))

        def concluir(confirmar):
            try:
                top.unbind("<MouseWheel>")
            except tk.TclError:
                pass
            if confirmar:
                usados = set()
                for pe, tv, mapa in escolhas:
                    pd = mapa.get((tv.selection() or [None])[0])
                    if pd is None or id(pd) in usados or pd["used_by"] is not None:
                        continue        # mesmo PDF escolhido 2x: vale a 1ª escolha
                    usados.add(id(pd))
                    pd["used_by"] = pe["paidId"]
                    pe["match"] = {"pdf": pd, "ocnf": False, "cc": False,
                                   "date": False, "docnum": False, "score": 0}
                    pe["pdf"] = pd["fn"]
                    pe["motivo"] = "escolhido por você"
                    pe["status"] = "CERTEZA"
            top.destroy()
            ev.set()

        b_ok = widgets.Botao(rodape, "✔  Confirmar escolhas", papel="acao",
                             command=lambda: concluir(True))
        b_ok.pack(side="left")
        widgets.Botao(rodape, "Deixar todas em dúvida", papel="neutro",
                      command=lambda: concluir(False)).pack(side="left", padx=10)
        top.protocol("WM_DELETE_WINDOW", lambda: concluir(False))
        top.grab_set()

    # ---------------------------------------------------------------- etapa 2
    def conectar(self):
        ini, fim = _data_api(self.v_ini.get()), _data_api(self.v_fim.get())
        if not ini or not fim:
            messagebox.showerror("Erro", "Datas inválidas. Use dd/mm/aaaa."); return
        if self.v_modo.get() == "auto" and not Path(self.v_pasta.get() or "").is_dir():
            messagebox.showerror("Erro", "Selecione a pasta dos PDFs renomeados."); return
        self.b1.config(state="disabled")
        self.log.delete("1.0", "end"); self._ph = False
        self.lbl.config(text="Conectando...")
        self.pb.config(mode="indeterminate")
        self.pb.start(12)
        self.worker = self.submeter("Anexar — carregar contas",
                                    self._t_conectar, ini, fim)

    def _t_conectar(self, ini, fim):
        inicio = time.time()
        self._log(f"⏱ Etapa 2 — início: {time.strftime('%H:%M:%S')}")

        def st(msg):                      # atualiza o texto de status da janela
            self.q.put(("status", msg))
        try:
            if self.mc is None:
                st("Abrindo o Chrome — faça o login se for pedido...")
            self.garantir_sessao()
            st("Capturando credenciais da tela de Pagamentos...")
            if not self.api.capturar_credenciais(self._log):
                raise RuntimeError("Não capturei a lista de pagamentos.")
            st("Buscando títulos pagos do período no servidor...")
            self._log(f"Buscando títulos PAGOS de {ini} a {fim} (todas as contas)...")

            def log_pg(m):
                self._log(m)
                st("Buscando títulos pagos — " + m.strip(" .") + "...")
            lanc = self.api.listar_pagos(ini, fim, log_pg)
            st("Organizando as contas bancárias...")
            self.pagos = mc_api.montar_pagos(lanc)
            contas = sorted({p["conta"] for p in self.pagos if p["conta"]})
            self._log(f"{len(lanc)} lançamento(s), {len(self.pagos)} pagamento(s), "
                      f"{len(contas)} conta(s) encontradas.")
            self._log(f"⏱ Etapa 2 — fim: {time.strftime('%H:%M:%S')} "
                      f"({_fmt_dur(time.time() - inicio)})")
            self.q.put(("contas", contas))
        except Exception as e:
            self._log(_texto_do_erro(e))
            self.q.put(("reabilitar", None))

    def _montar_contas(self, contas):
        for w in self.contas_box.winfo_children():
            w.destroy()
        self.vars_contas = {}
        cont = {c: 0 for c in contas}
        for p in self.pagos:
            if p["conta"] in cont:
                cont[p["conta"]] += 1
        colunas = 3
        for i, c in enumerate(contas):
            v = tk.BooleanVar(value=True)
            self.vars_contas[c] = v
            ttk.Checkbutton(self.contas_box, text=f"{c}  ({cont[c]})", variable=v
                            ).grid(row=i // colunas, column=i % colunas, sticky="w", padx=4)
        # "Marcar/Desmarcar todas" e a contagem sobem para o cabeçalho do
        # cartão: dentro da grade de contas eles disputavam a última linha com
        # as próprias contas, e mudavam de lugar conforme quantas havia.
        self.rodape_contas.limpar_links()
        self.rodape_contas.link("Marcar todas", lambda: self._todas_contas(True))
        self.rodape_contas.link("Desmarcar todas",
                                lambda: self._todas_contas(False))
        self._cont_por_conta = cont
        for v in self.vars_contas.values():
            v.trace_add("write", lambda *_a: self._contar_contas())
        self._contar_contas()

    def _todas_contas(self, marcar: bool):
        for v in self.vars_contas.values():
            v.set(marcar)

    def _contar_contas(self):
        try:
            marcadas = [c for c, v in self.vars_contas.items() if v.get()]
        except tk.TclError:
            return                       # aba fechando com o trace pendente
        pagos = sum(self._cont_por_conta.get(c, 0) for c in marcadas)
        self.rodape_contas.definir(
            texto=f"{len(marcadas)} conta(s) · {pagos} pagamento(s)"
                  + (f" · {len(self.vars_contas) - len(marcadas)} ficam de fora"
                     if len(marcadas) < len(self.vars_contas) else ""))

    # ---------------------------------------------------------------- etapa 2
    def executar(self):
        if self.worker and not self.worker.done():
            return
        # TUDO que vem do formulário é lido AQUI, na thread da interface, e vai
        # por argumento para o worker. Ler `StringVar`/`BooleanVar` é falar com
        # o Tcl, e o Tcl é de quem criou a janela: da thread do navegador isso
        # trava ou erra sem hora marcada — falha intermitente, que é justamente
        # a que não aparece em teste. Vale para as caixas de conta também.
        pasta = self.v_pasta.get()
        simular = self.v_dry.get()
        if self.v_modo.get() == "lista":
            lista = self.v_lista.get()
            if not Path(lista or "").exists():
                messagebox.showerror("Erro", "Selecione a lista (.csv/.xlsx)."); return
            alvo, args = self._t_lista, (pasta, lista, simular)
        else:
            if not self.pagos:
                messagebox.showerror("Erro", "Primeiro clique em \"1. Carregar contas\"."); return
            contas_sel = {c for c, v in self.vars_contas.items() if v.get()}
            termos = []
            if self.v_ign.get():
                termos += config.IGNORAR_TARIFAS
            if self.v_ign_ap.get():
                termos += config.IGNORAR_APORTES
            alvo, args = self._t_auto, (contas_sel, termos, pasta, simular)
        self._parar.clear()
        self._pausa.clear()
        self.b_pause.config(text="⏸ Pausar", state="normal")
        self.b_stop.config(state="normal")
        self.b2.config(state="disabled")
        self.worker = self.submeter("Anexar — casar e anexar", alvo, *args)

    def _t_auto(self, contas_sel: set, termos: list, pasta_pdfs: str,
                simular: bool):
        inicio = time.time()
        self._log(f"⏱ Etapa 3 — início: {time.strftime('%H:%M:%S')}")
        try:
            if not contas_sel:
                self._log("[!] Nenhuma conta marcada."); self.q.put(("reabilitar2", None)); return
            pagos = [p for p in self.pagos if p["conta"] in contas_sel]
            if termos:
                antes = len(pagos)
                pagos = [p for p in pagos
                         if not any(t in (_norm(p["desc"]) + " | " + _norm(p["categoria"]))
                                    for t in termos)]
                self._log(f"Ignorados por tipo: {antes - len(pagos)}")
            self._log(f"{len(pagos)} pagamento(s) nas contas marcadas. Verificando anexos...")

            if pagos and not self.api.capturar_credenciais_anexos(pagos[0]["launchId"]):
                raise RuntimeError("Não capturei as credenciais de anexos.")
            self.q.put(("max", len(pagos)))
            att = self.api.verificar_anexos([p["paidId"] for p in pagos], self._log,
                                            progresso=lambda i, n: self.q.put(("prog_verif", (i, n))),
                                            cancelar=self._checar_pausa)
            if self._parar.is_set():
                self._log("⏹ Interrompido pelo usuário durante a verificação.")
                self.q.put(("reabilitar2", None))
                return
            estados = {p["paidId"]: mc_api.estado_anexo(att, p["paidId"])
                       for p in pagos}
            com = [p for p in pagos
                   if estados[p["paidId"]] == mc_api.COM_ANEXO]
            sem = [p for p in pagos
                   if estados[p["paidId"]] == mc_api.SEM_ANEXO]
            nao_verif = [p for p in pagos
                         if estados[p["paidId"]] == mc_api.NAO_VERIFICADO]
            # Não verificado entra como PENDENTE: pular é assumir que já tem
            # comprovante sem nunca ter olhado. Tentar de novo, no pior caso,
            # devolve 'ja_tinha' — barato perto de deixar sem anexo.
            pendentes = sem + nao_verif
            self._log(f"Com comprovante: {len(com)} | SEM comprovante: {len(sem)}")
            if nao_verif:
                self._log(f"[aviso] {len(nao_verif)} pagamento(s) NÃO VERIFICADOS "
                          "(a consulta de anexos falhou) — vão junto dos "
                          "pendentes por precaução.")
            self._log(f"⏱ Verificação de anexos: {_fmt_dur(time.time() - inicio)}")
            ini_anexar = time.time()

            pdfs = matcher.carregar_pdfs(Path(pasta_pdfs), self._log)
            self._log(f"{len(pdfs)} PDF(s) válidos na pasta.")
            certezas, duvidas, sem_par = matcher.casar(pendentes, pdfs)
            self._log(f"Casamentos com certeza: {len(certezas)} | dúvida: {len(duvidas)} "
                      f"| sem par: {len(sem_par)}\n")

            if duvidas and not self._parar.is_set():
                self._log("Abrindo a janela para você resolver as dúvidas...")
                ev = Event()
                self._esperas.add(ev)       # fechar() precisa soltar isto
                try:
                    self.q.put(("duvidas", (duvidas, ev)))
                    while not ev.wait(timeout=0.5):
                        if self._parar.is_set():
                            break
                finally:
                    self._esperas.discard(ev)
                resolvidas = [p for p in duvidas if p["status"] == "CERTEZA"]
                if resolvidas:
                    self._log(f"{len(resolvidas)} dúvida(s) resolvida(s) por você.\n")
                    certezas += resolvidas
                    duvidas = [p for p in duvidas if p["status"] == "DUVIDA"]

            resultados = []
            self.q.put(("max", len(certezas)))
            pasta = Path(pasta_pdfs)
            for i, pe in enumerate(certezas, 1):
                if self._checar_pausa():
                    self._log("⏹ Interrompido pelo usuário — gerando relatório "
                              "com o que já foi feito.")
                    break
                arq = pasta / pe["pdf"]
                vals = [_fmt_val(v) for v in pe.get("valores", [pe["valor"]])]
                r = self.mc.anexar(pe["launchId"], _fmt_val(pe["valor"]), arq,
                                   doc=pe["doc"] or None, dry_run=simular,
                                   valores=vals)
                if r.startswith("erro:"):
                    self._log(f"   ({r}) — recarregando o sistema e tentando de novo...")
                    self.mc.resetar()
                    r = self.mc.anexar(pe["launchId"], _fmt_val(pe["valor"]), arq,
                                       doc=pe["doc"] or None, dry_run=simular,
                                       valores=vals)
                pe["resultado"] = r
                resultados.append(pe)
                self.q.put(("prog", (i, sum(1 for x in resultados if not x["resultado"].startswith("erro")), 0)))
                self._log(f"[{i}/{len(certezas)}] {_fmt_val(pe['valor'])}  {pe['pdf']}  -> {r}")

            saida = self._relatorio(resultados, duvidas, sem_par, pasta_pdfs)
            ok = sum(1 for x in resultados
                     if x["resultado"] in ("anexado", "anexado_sem_tag", "ja_tinha", "dry_run"))
            self._log(f"⏱ Anexos: {_fmt_dur(time.time() - ini_anexar)}")
            self._log(f"\nConcluído. Anexados/ok: {ok} de {len(certezas)}. Relatório: {saida}")
            self._log(f"⏱ Etapa 3 — fim: {time.strftime('%H:%M:%S')} "
                      f"(total: {_fmt_dur(time.time() - inicio)})")
            self.q.put(("fim", (ok, len(certezas), len(duvidas), len(sem_par), saida)))
        except Exception as e:
            self._log(_texto_do_erro(e))
            self.q.put(("reabilitar2", None))

    def _t_lista(self, pasta_pdfs: str, lista: str, simular: bool):
        inicio = time.time()
        self._log(f"⏱ Início: {time.strftime('%H:%M:%S')}")
        try:
            pasta_pdfs = (pasta_pdfs or "").strip() or None
            tarefas = planilha.carregar_tarefas(Path(lista),
                                                pasta_pdfs=pasta_pdfs)
            self._log(f"{len(tarefas)} linha(s) na lista. Simular={simular}")
            self.garantir_sessao()
            self.q.put(("max", len(tarefas)))
            ok = 0
            for i, t in enumerate(tarefas, 1):
                if self._checar_pausa():
                    self._log("⏹ Interrompido pelo usuário.")
                    break
                if not t["launchId"]:
                    r = "erro:sem_link_do_lancamento"
                elif not t["valor"]:
                    r = "erro:sem_valor"
                elif t["arquivo"] is None:
                    r = "erro:pdf_nao_encontrado_na_pasta"
                else:
                    r = self.mc.anexar(t["launchId"], t["valor"], t["arquivo"],
                                       doc=t.get("doc") or None, dry_run=simular)
                    if r.startswith("erro:"):
                        self._log(f"   ({r}) — recarregando o sistema e tentando de novo...")
                        self.mc.resetar()
                        r = self.mc.anexar(t["launchId"], t["valor"], t["arquivo"],
                                           doc=t.get("doc") or None,
                                           dry_run=simular)
                if r in ("anexado", "anexado_sem_tag", "ja_tinha", "dry_run"):
                    ok += 1
                self.q.put(("prog", (i, ok, i - ok)))
                self._log(f"[{i}/{len(tarefas)}] {t['valor']}  {t['arquivo_bruto']}  -> {r}")
            self._log(f"\nConcluído: {ok}/{len(tarefas)} ok.")
            self._log(f"⏱ Fim: {time.strftime('%H:%M:%S')} — tempo total: "
                      f"{_fmt_dur(time.time() - inicio)}")
            self.q.put(("fim", (ok, len(tarefas), 0, 0, "")))
        except Exception as e:
            self._log(_texto_do_erro(e))
            self.q.put(("reabilitar2", None))

    # ---------------------------------------------------------------- saída
    def _relatorio(self, anexados, duvidas, sem_par, pasta_pdfs: str = "") -> str:
        """Escreve o Excel ao lado dos PDFs. Roda na thread do navegador, então
        a pasta chega por ARGUMENTO — lida do `v_pasta` daqui, seria tkinter
        fora da thread da interface."""
        wb = Workbook(); wb.remove(wb.active)
        verde = PatternFill("solid", fgColor="1B7837")
        branco = Font(bold=True, color="FFFFFF")
        H = ["Valor", "Data", "Favorecido", "Centro de custo", "Conta",
             "Descrição", "Nº doc", "OC/NF", "PDF", "Motivo/Candidatos",
             "Resultado", "Link"]

        def aba(nome, linhas):
            ws = wb.create_sheet(nome)
            for j, h in enumerate(H, 1):
                c = ws.cell(1, j, h); c.font = branco; c.fill = verde
            for i, r in enumerate(linhas, 2):
                for j, v in enumerate(r, 1):
                    ws.cell(i, j, v)
            for col, w in zip("ABCDEFGHIJKL",
                              [11, 9, 30, 32, 26, 38, 16, 14, 45, 30, 16, 58]):
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"

        def comuns(p):
            return [_fmt_val(p["valor"]), p["dataFull"], p.get("favorecido", ""),
                    "; ".join(p["works"]), p["conta"], p["desc"], p["doc"],
                    ", ".join(p.get("ocs") or [])]

        aba("ANEXADOS", [comuns(p) + [p["pdf"], p["motivo"],
                                      p.get("resultado", ""), LINK + p["launchId"]]
                         for p in anexados])
        aba("DUVIDA", [comuns(p) + ["", _resumo_cands(p), "", LINK + p["launchId"]]
                       for p in duvidas])
        aba("SEM PAR", [comuns(p) + ["", "", "", LINK + p["launchId"]]
                        for p in sem_par])
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = Path(pasta_pdfs or ".") / f"relatorio_anexos_{stamp}.xlsx"
        wb.save(out)
        return str(out).replace("\\", "/")

    # ---------------------------------------------------------------- UI pump
    def _drain(self):
        try:
            while True:
                kind, val = self.q.get_nowait()
                if kind == "log":
                    if getattr(self, "_ph", False):    # limpa o estado inicial
                        self.log.delete("1.0", "end"); self._ph = False
                    self.log.insert("end", val + "\n"); self.log.see("end")
                elif kind == "status":
                    self.lbl.config(text=val)
                elif kind == "max":
                    self.pb.stop()
                    self.pb.config(mode="determinate", maximum=max(val, 1), value=0)
                elif kind == "prog":
                    i, ok, err = val
                    self.barra_exec.progresso(i, int(self.pb.cget("maximum")))
                    self.lbl.config(text=f"{i} processados — {ok} ok" + (f", {err} erros" if err else ""))
                elif kind == "prog_verif":
                    i, n = val
                    self.barra_exec.progresso(i, n)
                    self.lbl.config(text=f"Verificando comprovantes já anexados: {i}/{n}")
                elif kind == "contas":
                    self.pb.stop()
                    self.pb.config(mode="determinate", value=0)
                    self._montar_contas(val)
                    self.b1.config(state="normal")
                    self.b2.config(state="normal")
                    self.lbl.config(text="Contas carregadas. Marque as "
                                         "desejadas e clique em Casar e anexar.")
                elif kind == "reabilitar0":
                    self.b0.config(state="normal")
                elif kind == "reabilitar":
                    self.pb.stop()
                    self.pb.config(mode="determinate", value=0)
                    self.b1.config(state="normal")
                    self.lbl.config(text="Erro ao carregar — veja o Registro.")
                elif kind == "reabilitar2":
                    self.b2.config(state="normal")
                    self.b_pause.config(text="⏸ Pausar", state="disabled")
                    self.b_stop.config(state="disabled")
                elif kind == "duvidas":
                    self._janela_duvidas(*val)
                elif kind == "fim":
                    ok, tot, duv, sp, saida = val
                    self.b2.config(state="normal")
                    self.b_pause.config(text="⏸ Pausar", state="disabled")
                    self.b_stop.config(state="disabled")
                    if saida:
                        self.ultimo_relatorio = saida
                        self.b_rel.config(state="normal")
                    self.barra_exec.terminou(
                        f"Concluído: {ok}/{tot} ok"
                        + (f" | {duv} dúvidas, {sp} sem par" if duv or sp else ""))
                    # O "Anexados no mês" do Início sai daqui: quem contou foi
                    # esta rotina, e recontar custaria outra sessão do ERP.
                    widgets.registrar_atividade(
                        "anx", "Casar e anexar",
                        "atencao" if (duv or sp) else "ok",
                        f"{ok} de {tot} anexado(s)"
                        + (f" · {duv} em dúvida, {sp} sem par" if duv or sp
                           else ""),
                        {"anexados": ok, "pagos": tot, "duvidas": duv,
                         "sem_par": sp})
                    msg = f"Anexados/ok: {ok} de {tot}"
                    if duv or sp:
                        msg += f"\nDúvidas: {duv}\nSem par: {sp}"
                    if saida:
                        msg += f"\n\nRelatório:\n{saida}"
                    messagebox.showinfo("Concluído", msg)
        except queue.Empty:
            pass
        except Exception as e:                              # noqa: BLE001
            # A bomba de UI NUNCA pode morrer: sem ela o log para de aparecer,
            # os botões nunca voltam e a aba parece travada — enquanto a thread
            # do navegador segue trabalhando. Registrar e continuar drenando é
            # sempre melhor do que deixar a interface muda.
            config.diag(f"_drain (Anexar) falhou: {e!r}")
        finally:
            self.after(150, self._drain)

    def fechar(self):
        """Fecha o navegador e a thread de trabalho (chamar ao sair do app).

        A ORDEM importa. A thread do navegador pode estar parada esperando
        gente: pausada, ou dentro da janela de dúvidas (que bloqueia em
        `ev.wait()` até alguém responder). Fechar o Chrome primeiro deixaria
        essa thread presa para sempre e o processo não morreria — a janela
        some e o app fica de fundo, segurando o perfil do Chrome."""
        # 1) solta quem está esperando
        try:
            self._parar.set()
            self._pausa.clear()
        except Exception:
            pass
        for ev in list(getattr(self, "_esperas", ())):
            try:
                ev.set()
            except Exception:
                pass
        for filho in list(self.winfo_children()):
            if isinstance(filho, tk.Toplevel):
                try:
                    filho.grab_release()
                    filho.destroy()
                except Exception:
                    pass

        # 2) fecha o navegador na thread dele (exigência do Playwright)
        try:
            if self.mc:
                self.exec.submit(self.mc.__exit__, None, None, None).result(timeout=8)
                self.mc = None
        except Exception as e:
            config.diag(f"fechar(): o Chrome não fechou limpo: {e!r}")
        # 3) último recurso: descarta o que ainda estiver na fila
        try:
            self.exec.shutdown(wait=False, cancel_futures=True)
        except TypeError:                    # Python < 3.9
            self.exec.shutdown(wait=False)
        except Exception:
            pass


def main():
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)   # texto nítido em telas HiDPI
    except Exception:
        pass
    root = tk.Tk()
    root.title("Anexar Comprovantes — Mais Controle")
    try:
        root.state("zoomed")            # janela ocupando a tela (Windows)
    except tk.TclError:
        root.geometry("1100x720")
    try:
        import sv_ttk                   # tema moderno (visual Windows 11)
        sv_ttk.set_theme("light")
    except Exception:
        pass
    app = AnexarFrame(root)
    app.pack(fill="both", expand=True)

    def _sair():
        app.fechar()
        root.destroy()
    root.protocol("WM_DELETE_WINDOW", _sair)
    root.mainloop()


if __name__ == "__main__":
    main()
