# -*- coding: utf-8 -*-
"""
Conferência pós-anexo.

Varre os títulos PAGOS do período e lista quem ainda está SEM comprovante.
Opcionalmente, abre cada anexo existente e confere se o VALOR (e a data)
do PDF batem com o lançamento — divergências vão para o relatório.

Compartilha o navegador da tela "Anexar Comprovantes" (mesma sessão logada
e mesma thread de trabalho).
"""
import io
import queue
import time
from threading import Event
from datetime import date, datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from . import config, mc_api
except ImportError:
    import config, mc_api

try:                                     # utilitários compartilhados (raiz)
    import util
except ModuleNotFoundError:              # rodando este módulo isoladamente
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    import util

try:                                     # widgets compartilhados (raiz)
    import widgets
except ModuleNotFoundError:              # rodando este módulo isoladamente
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    import widgets

CampoData = widgets.CampoData
# Só o _texto_do_erro segue vindo do Anexar: ele conhece o SemRede do
# mc_client, que é daquela aba. O campo de data e os formatos são de todos.
from anexar_comprovantes import _texto_do_erro

LINK = config.MC_URL_LANCAMENTO
_fmt_dur = util.fmt_dur
_fmt_val = util.fmt_val
_data_api = util.data_api
_norm = util.norm


def _texto_pdf(dados: bytes) -> str:
    """Extrai o texto de um PDF baixado; usa OCR se não houver texto."""
    try:
        import pdfplumber
    except ImportError:
        config.diag("conferência: pdfplumber indisponível, não dá para "
                    "conferir o conteúdo dos anexos")
        return ""
    try:
        with pdfplumber.open(io.BytesIO(dados)) as pl:
            txt = "\n".join((pg.extract_text() or "") for pg in pl.pages)
        if len(txt.strip()) >= 30:
            return txt
        try:                                # sem camada de texto -> OCR
            from separar_renomear import _ocr_pagina
            with pdfplumber.open(io.BytesIO(dados)) as pl:
                return "\n".join(_ocr_pagina(pg, lambda m: None)
                                 for pg in pl.pages)
        except Exception as e:
            config.diag(f"conferência: OCR do anexo falhou ({e!r})")
            return txt
    except Exception as e:
        config.diag(f"conferência: não consegui ler o PDF do anexo ({e!r})")
        return ""


def _valor_no_texto(cents: int, txt: str) -> bool:
    inteiro, cent = divmod(int(cents), 100)
    com_milhar = f"{inteiro:,}".replace(",", ".")
    formas = {f"{com_milhar},{cent:02d}", f"{inteiro},{cent:02d}"}
    return any(f in txt for f in formas)


def _host_path(url) -> str:
    if not url:
        return "(nenhuma URL no registro do anexo)"
    try:
        from urllib.parse import urlsplit
        s = urlsplit(url)
        return f"{s.scheme}://{s.netloc}{s.path}"
    except Exception:
        return "(url ilegível)"


def _estrutura_anexo(item, prof=0) -> str:
    """Resumo SEGURO da estrutura de um registro de anexo: nomes dos campos e
    tipos (e host/caminho de URLs), sem os valores sensíveis. Serve para
    descobrir como baixar o arquivo."""
    try:
        if isinstance(item, dict):
            return "{" + ", ".join(f"{k}: {_estrutura_anexo(v, prof + 1)}"
                                   for k, v in list(item.items())[:25]) + "}"
        if isinstance(item, list):
            return "[" + (_estrutura_anexo(item[0], prof + 1) if item else "") + "]"
        if isinstance(item, bool):
            return "bool"
        if isinstance(item, (int, float)):
            return "num"
        if item is None:
            return "null"
        if isinstance(item, str):
            return "<url " + _host_path(item) + ">" if item.startswith("http") \
                else f"str({len(item)})"
        return type(item).__name__
    except Exception:
        return "?"


class ConferenciaFrame(ttk.Frame):
    """Tela de conferência (usa a sessão da tela Anexar)."""

    def __init__(self, master, anexar_frame):
        super().__init__(master)
        self.anx = anexar_frame
        self.q = queue.Queue()
        self.ultimo_relatorio = None
        self.worker = None
        # A conferência com conteúdo baixa e lê um PDF por pagamento: em mês
        # cheio passa de meia hora. Sem Parar, a única saída era matar o app —
        # e matar o app no meio deixa o Chrome órfão segurando o perfil.
        self._parar = Event()
        hoje = date.today()
        self.v_ini = tk.StringVar(value=hoje.replace(day=1).strftime("%d/%m/%Y"))
        self.v_fim = tk.StringVar(value=hoje.strftime("%d/%m/%Y"))
        self.v_ign = tk.BooleanVar(value=True)
        self.v_ign_ap = tk.BooleanVar(value=True)
        self.v_conteudo = tk.BooleanVar(value=False)
        self._montar()
        try:                             # já nasce na cor do tema (sem flash)
            self.aplicar_cores(util.cor_escura(ttk.Style().lookup("TFrame", "background")))
        except Exception:
            pass
        self.after(150, self._drain)

    def _montar(self):
        PADX = 14

        # ---- cabeçalho
        cab = ttk.Frame(self)
        cab.pack(fill="x", padx=PADX, pady=(12, 4))
        ttk.Label(cab, text="Conferência",
                  font=("Segoe UI", 14, "bold")).pack(anchor="w")
        self.lbl_sub = ttk.Label(
            cab, foreground="#6b6b6b",
            text="Lista os pagos sem comprovante e, opcionalmente, confere o "
                 "conteúdo dos anexos existentes.")
        self.lbl_sub.pack(anchor="w")

        # ---- card: período
        f1 = ttk.LabelFrame(self, text=" Período da conferência ",
                            padding=(12, 8, 12, 10))
        f1.pack(fill="x", padx=PADX, pady=6)
        ttk.Label(f1, text="Data de pagamento — de:").grid(row=0, column=0, sticky="w", pady=4)
        CampoData(f1, self.v_ini).grid(row=0, column=1, sticky="w", padx=(6, 14))
        ttk.Label(f1, text="até:").grid(row=0, column=2, sticky="e")
        CampoData(f1, self.v_fim).grid(row=0, column=3, sticky="w", padx=(6, 14))
        ttk.Label(f1, text="(dd/mm/aaaa)").grid(row=0, column=4, sticky="w")
        ttk.Checkbutton(f1, text="Ignorar tarifas bancárias, IOF, cesta e pacote de serviços",
                        variable=self.v_ign).grid(row=1, column=0, columnspan=5, sticky="w", pady=(6, 0))
        ttk.Checkbutton(f1, text="Ignorar aportes de capital e distribuição de lucros",
                        variable=self.v_ign_ap).grid(row=2, column=0, columnspan=5, sticky="w")
        ttk.Checkbutton(f1, text="Conferir também o CONTEÚDO dos anexos "
                                 "(abre cada PDF e checa valor e data — mais demorado)",
                        variable=self.v_conteudo
                        ).grid(row=3, column=0, columnspan=5, sticky="w", pady=(0, 2))

        # ---- barra de ação (fixa no rodapé)
        acao = ttk.Frame(self)
        acao.pack(side="bottom", fill="x", padx=PADX, pady=(6, 12))
        self.btn = ttk.Button(acao, text="▶  Conferir anexos do período",
                              command=self._executar)
        self.btn.pack(side="right", ipadx=10)
        try:
            self.btn.configure(style="Accent.TButton")
        except tk.TclError:
            pass
        self.b_stop = ttk.Button(acao, text="⏹  Parar",
                                 command=self._parar_click, state="disabled")
        self.b_stop.pack(side="right", padx=(0, 8))
        self.b_rel = ttk.Button(acao, text="📄  Abrir relatório",
                                command=self._abrir_relatorio, state="disabled")
        self.b_rel.pack(side="right", padx=(0, 8))
        self.lbl = ttk.Label(acao, text="Pronto.")
        self.lbl.pack(side="left")
        self.pb = ttk.Progressbar(acao, mode="determinate")
        self.pb.pack(side="left", fill="x", expand=True, padx=12)

        # ---- card: registro (ocupa o espaço restante)
        reg = ttk.LabelFrame(self, text=" Registro ", padding=(10, 6, 10, 10))
        reg.pack(fill="both", expand=True, padx=PADX, pady=6)
        self.log = tk.Text(reg, wrap="word", relief="flat", borderwidth=0,
                           highlightthickness=0, background="#ffffff",
                           font=("Consolas", 10))
        self.log.pack(fill="both", expand=True)
        self.log.tag_configure("ph", justify="center", foreground="#8a8a8a",
                               spacing1=6, font=("Segoe UI", 11))
        self._mostrar_placeholder()

    def _mostrar_placeholder(self):
        self.log.delete("1.0", "end")
        self.log.insert("end", "\n\n", "ph")
        self.log.insert("end", "O resultado da conferência aparecerá aqui.\n", "ph")
        self.log.insert("end", "\nInforme o período e clique em "
                               "“Conferir anexos do período”.\n", "ph")

    def aplicar_cores(self, escuro: bool):
        if escuro:
            self.log.configure(background="#252525", foreground="#e6e6e6",
                               insertbackground="#e6e6e6")
            muted = "#9a9a9a"
        else:
            self.log.configure(background="#ffffff", foreground="#000000",
                               insertbackground="#000000")
            muted = "#5f5f5f"
        self.log.tag_configure("ph", foreground="#8a8a8a")
        try:
            self.lbl_sub.configure(foreground=muted)
        except Exception:
            pass

    def _log(self, m):
        self.q.put(("log", m))

    def _drain(self):
        try:
            while True:
                kind, val = self.q.get_nowait()
                if kind == "log":
                    self.log.insert("end", val + "\n"); self.log.see("end")
                elif kind == "status":
                    self.lbl.config(text=val)
                elif kind == "max":
                    self.pb.config(maximum=max(val, 1), value=0)
                elif kind == "prog":
                    self.pb.config(value=val)
                elif kind == "fim":
                    self.btn.config(state="normal")
                    self.b_stop.config(state="disabled")
                    self.pb.config(value=0)
                    if val:
                        self.ultimo_relatorio = val
                        self.b_rel.config(state="normal")
        except queue.Empty:
            pass
        except Exception as e:                              # noqa: BLE001
            # Ver o comentário gêmeo em anexar_comprovantes._drain: a bomba de
            # UI morrendo deixa a aba muda com a thread ainda trabalhando.
            config.diag(f"_drain (Conferência) falhou: {e!r}")
        finally:
            self.after(150, self._drain)

    def _parar_click(self):
        self._parar.set()
        self._log("\n⏹ Parando… termino o item atual e gero o relatório com o "
                  "que já foi conferido.")
        self.b_stop.config(state="disabled")

    def _abrir_relatorio(self):
        import os
        if self.ultimo_relatorio and Path(self.ultimo_relatorio).exists():
            try:
                os.startfile(self.ultimo_relatorio)
            except OSError as e:
                messagebox.showerror("Erro", f"Não consegui abrir o relatório:\n{e}")

    # ------------------------------------------------------------ execução
    def _executar(self):
        ini, fim = _data_api(self.v_ini.get()), _data_api(self.v_fim.get())
        if not ini or not fim:
            messagebox.showerror("Erro", "Datas inválidas. Use dd/mm/aaaa.")
            return
        if self.anx.avisar_se_ocupado("a Conferência"):
            return
        self._parar.clear()
        self.btn.config(state="disabled")
        self.b_stop.config(state="normal")
        self.log.delete("1.0", "end")
        self.lbl.config(text="Conferindo...")
        self.worker = self.anx.submeter("Conferência", self._t_conferir,
                                        ini, fim)

    def _t_conferir(self, ini, fim):
        inicio = time.time()
        self._log(f"⏱ Início: {time.strftime('%H:%M:%S')}")
        try:
            api = self.anx.garantir_sessao(self._log)
            if not api.capturar_credenciais(self._log):
                raise RuntimeError("Não capturei a lista de pagamentos.")
            self.q.put(("status", "Buscando títulos pagos do período..."))
            lanc = api.listar_pagos(ini, fim, self._log)
            pagos = mc_api.montar_pagos(lanc)
            termos = []
            if self.v_ign.get():
                termos += config.IGNORAR_TARIFAS
            if self.v_ign_ap.get():
                termos += config.IGNORAR_APORTES
            if termos:
                pagos = [p for p in pagos
                         if not any(t in (_norm(p["desc"]) + " | " + _norm(p["categoria"]))
                                    for t in termos)]
            self._log(f"{len(pagos)} pagamento(s) no período (após filtros).")
            from collections import Counter
            contas = Counter(p["conta"] for p in pagos if p["conta"])
            if contas:
                self._log("Por conta: " + " | ".join(f"{c}: {n}"
                                                      for c, n in contas.items()))
            if pagos and not api.capturar_credenciais_anexos(pagos[0]["launchId"]):
                raise RuntimeError("Não capturei as credenciais de anexos.")

            self.q.put(("status", "Verificando quem tem anexo..."))
            self.q.put(("max", len(pagos)))
            att = api.verificar_anexos([p["paidId"] for p in pagos], self._log,
                                       progresso=lambda i, n: self.q.put(("prog", i)),
                                       cancelar=self._parar.is_set)
            estados = {p["paidId"]: mc_api.estado_anexo(att, p["paidId"])
                       for p in pagos}
            sem = [p for p in pagos if estados[p["paidId"]] == mc_api.SEM_ANEXO]
            com = [p for p in pagos if estados[p["paidId"]] == mc_api.COM_ANEXO]
            # Nem "com" nem "sem": a consulta falhou. Antes caíam no balde do
            # "com anexo" e sumiam do relatório — uma conferência que dizia
            # "tudo certo" justamente sobre o que não foi conferido.
            nao_verif = [p for p in pagos
                         if estados[p["paidId"]] == mc_api.NAO_VERIFICADO]
            self._log(f"\nCom anexo: {len(com)} | SEM anexo: {len(sem)}"
                      + (f" | NÃO VERIFICADOS: {len(nao_verif)}" if nao_verif else ""))
            for p in sem:
                self._log(f"  SEM ANEXO: {_fmt_val(p['valor'])}  {p['dataFull']}  "
                          f"{p['conta']}  {p.get('favorecido') or '—'}  "
                          f"{p['desc'][:60]}")
            for p in nao_verif:
                p["confere"] = "não verificado (a consulta de anexos falhou)"
                self._log(f"  NÃO VERIFICADO: {_fmt_val(p['valor'])}  "
                          f"{p['dataFull']}  {p['conta']}  {p['desc'][:60]}")

            divergentes, conferidos, nao_conferiveis = [], [], []
            if self.v_conteudo.get() and com:
                self.q.put(("status", "Conferindo o conteúdo dos anexos..."))
                self._log(f"\nConferindo o conteúdo de {len(com)} anexo(s)...")
                self.q.put(("max", len(com)))
                for i, p in enumerate(com, 1):
                    if self._parar.is_set():
                        self._log(f"⏹ Interrompido: {i - 1} de {len(com)} "
                                  "anexo(s) conferido(s). O relatório sai com "
                                  "o que já foi feito.")
                        break
                    self.q.put(("prog", i))
                    try:
                        itens = api.listar_anexos(p["paidId"])
                        baixados = 0
                        textos = []
                        for item in itens:
                            url = mc_api.achar_url_anexo(item)
                            dados = api.baixar_anexo(url) if url else None
                            txt_i = _texto_pdf(dados) if dados else ""
                            if not getattr(self, "_diag_feito", False):
                                self._diag_feito = True
                                self._log("[diag] estrutura do anexo: "
                                          + _estrutura_anexo(item))
                                self._log("[diag] URL: " + _host_path(url))
                                self._log("[diag] download: "
                                          + (f"OK ({len(dados)} bytes)" if dados else "FALHOU"))
                                self._log(f"[diag] texto/OCR: {len(txt_i.strip())} chars")
                            if dados:
                                baixados += 1
                                textos.append(txt_i)
                        texto = "\n".join(t for t in textos if t)
                        if not itens:
                            p["confere"] = "não conferível (anexo não listado pela API)"
                            nao_conferiveis.append(p)
                            continue
                        if baixados == 0:
                            p["confere"] = "não conferível (não consegui baixar o arquivo)"
                            nao_conferiveis.append(p)
                            continue
                        if not texto.strip():
                            p["confere"] = "não conferível (PDF sem texto e OCR não leu)"
                            nao_conferiveis.append(p)
                            continue
                        vals = p.get("valores") or [p["valor"]]
                        val_ok = any(_valor_no_texto(v, texto) for v in vals)
                        dt = p["dataFull"]
                        data_ok = bool(dt) and (
                            f"{dt[8:10]}/{dt[5:7]}/{dt[0:4]}" in texto)
                        if val_ok:
                            p["confere"] = "OK" + ("" if data_ok else " (data não encontrada)")
                            conferidos.append(p)
                        else:
                            p["confere"] = "DIVERGENTE (valor não encontrado no PDF)"
                            divergentes.append(p)
                            self._log(f"  DIVERGENTE: {_fmt_val(p['valor'])}  "
                                      f"{p['dataFull']}  {p['conta']}  {p['desc'][:50]}")
                    except Exception as e:
                        p["confere"] = f"erro: {str(e)[:60]}"
                        nao_conferiveis.append(p)
                self._log(f"Conteúdo: {len(conferidos)} ok | {len(divergentes)} "
                          f"divergente(s) | {len(nao_conferiveis)} não conferível(is)")
                if nao_conferiveis:
                    for m, n in Counter(p.get("confere", "?")
                                        for p in nao_conferiveis).items():
                        self._log(f"   • {n}× {m}")

            saida = self._relatorio(sem, divergentes, conferidos,
                                    nao_conferiveis, nao_verif)
            self._log(f"\nRelatório: {saida}")
            self._log(f"⏱ Fim: {time.strftime('%H:%M:%S')} — tempo total: "
                      f"{_fmt_dur(time.time() - inicio)}")
            self.q.put(("status",
                        f"Concluído: {len(sem)} sem anexo"
                        + (f", {len(divergentes)} divergente(s)"
                           if self.v_conteudo.get() else "")
                        + (f", {len(nao_verif)} não verificado(s)"
                           if nao_verif else "")))
            self.q.put(("fim", saida))
        except Exception as e:
            self._log(_texto_do_erro(e))
            self.q.put(("status", "Erro — veja o Registro."))
            self.q.put(("fim", None))

    def _relatorio(self, sem, divergentes, conferidos, nao_conf,
                   nao_verif=()) -> str:
        wb = Workbook(); wb.remove(wb.active)
        verde = PatternFill("solid", fgColor="1B7837")
        branco = Font(bold=True, color="FFFFFF")
        H = ["Valor", "Data", "Favorecido", "Centro de custo", "Conta",
             "Descrição", "Nº doc", "OC/NF", "Situação", "Link"]

        def aba(nome, linhas):
            ws = wb.create_sheet(nome)
            for j, h in enumerate(H, 1):
                c = ws.cell(1, j, h); c.font = branco; c.fill = verde
            for i, p in enumerate(linhas, 2):
                for j, v in enumerate(
                        [_fmt_val(p["valor"]), p["dataFull"],
                         p.get("favorecido", ""), "; ".join(p["works"]),
                         p["conta"], p["desc"], p["doc"],
                         ", ".join(p.get("ocs") or []), p.get("confere", ""),
                         LINK + str(p["launchId"])], 1):
                    ws.cell(i, j, v)
            for col, w in zip("ABCDEFGHIJ",
                              [11, 11, 30, 30, 26, 40, 14, 14, 30, 58]):
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"

        aba("SEM ANEXO", sem)
        # Aba própria, e nunca misturada com CONFERIDOS: estes pagamentos não
        # foram olhados. Enterrá-los junto do que passou é como dizer que
        # passaram.
        if nao_verif:
            aba("NAO VERIFICADOS", list(nao_verif))
        if divergentes or conferidos or nao_conf:
            aba("DIVERGENTES", divergentes)
            aba("CONFERIDOS", conferidos + nao_conf)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = _pasta_relatorio()
        out = base / f"relatorio_conferencia_{stamp}.xlsx"
        wb.save(out)
        return str(out).replace("\\", "/")


def _pasta_relatorio() -> Path:
    import sys
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent
