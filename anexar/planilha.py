# -*- coding: utf-8 -*-
"""
Carrega a lista de anexos a partir de:

  1) lista_para_anexar.csv  — colunas: launchId,valor,arquivo_pdf
     (arquivo_pdf pode ser caminho completo ou só o nome do arquivo)

  2) relacao_casamento.xlsx — aba "CERTEZA", colunas:
     valor | data | centro de custo | conta | descricao | doc | PDF(s) | obs | link
     (o launchId é extraído do link; o PDF é procurado na pasta informada)

Retorna lista de dicts: {launchId, valor, arquivo (Path), doc, origem}
"""
import csv
import re
from pathlib import Path

RE_LAUNCH = re.compile(r"payable-installments/([0-9a-fA-F-]{36})")


def _resolver_pdf(bruto: str, pastas: list[Path]) -> Path | None:
    """Resolve o caminho do PDF: absoluto, ou procura o nome nas pastas dadas."""
    bruto = (bruto or "").strip().strip('"')
    if not bruto:
        return None
    p = Path(bruto)
    if p.is_absolute() and p.exists():
        return p
    # nome do arquivo, aceitando separadores de Windows e de Unix
    nome = re.split(r"[\\/]", bruto)[-1].strip()
    for pasta in pastas:
        cand = pasta / nome
        if cand.exists():
            return cand
    return None


def _de_csv(caminho: Path, pastas: list[Path]) -> list[dict]:
    itens = []
    with open(caminho, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            arq = _resolver_pdf(row.get("arquivo_pdf", ""), pastas)
            itens.append({
                "launchId": (row.get("launchId") or "").strip(),
                "valor": (row.get("valor") or "").strip(),
                "arquivo": arq,
                "arquivo_bruto": (row.get("arquivo_pdf") or "").strip(),
                "doc": "",
                "origem": "csv",
            })
    return itens


def _de_xlsx(caminho: Path, pastas: list[Path]) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(caminho, data_only=True)
    nome_aba = next((n for n in wb.sheetnames if n.strip().upper() == "CERTEZA"),
                    wb.sheetnames[0])
    ws = wb[nome_aba]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    def col(*nomes):
        for n in nomes:
            for i, h in enumerate(headers):
                if h.startswith(n):
                    return i
        return None

    cv, cdoc, cpdf, clink = col("valor"), col("doc"), col("pdf"), col("link")
    itens = []
    for r in ws.iter_rows(min_row=2):
        if all(c.value in (None, "") for c in r):
            continue
        link = ""
        if clink is not None:
            cell = r[clink]
            link = (cell.hyperlink.target if cell.hyperlink else str(cell.value or "")).strip()
        m = RE_LAUNCH.search(link)
        launch = m.group(1) if m else ""
        nome_pdf = str(r[cpdf].value or "").strip() if cpdf is not None else ""
        valor = r[cv].value if cv is not None else ""
        if isinstance(valor, float) and valor == int(valor):
            valor = int(valor)
        itens.append({
            "launchId": launch,
            "valor": str(valor).strip(),
            "arquivo": _resolver_pdf(nome_pdf, pastas),
            "arquivo_bruto": nome_pdf,
            "doc": str(r[cdoc].value or "").strip() if cdoc is not None else "",
            "origem": f"xlsx:{nome_aba}",
        })
    return itens


def carregar_tarefas(caminho: Path, pasta_pdfs: Path | None = None) -> list[dict]:
    """Carrega o CSV ou XLSX. Procura PDFs: caminho absoluto da própria lista,
    depois na pasta_pdfs (se dada), depois na pasta onde a lista está."""
    caminho = Path(caminho)
    pastas = []
    if pasta_pdfs:
        pastas.append(Path(pasta_pdfs))
    pastas.append(caminho.parent)
    if caminho.suffix.lower() == ".csv":
        return _de_csv(caminho, pastas)
    if caminho.suffix.lower() in (".xlsx", ".xlsm"):
        return _de_xlsx(caminho, pastas)
    raise ValueError(f"Formato não suportado: {caminho.suffix} (use .csv ou .xlsx)")
