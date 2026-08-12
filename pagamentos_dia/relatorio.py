# -*- coding: utf-8 -*-
"""
Relatório dos pagamentos do dia: uma aba de Excel por conta bancária.

Só regra de negócio e Excel. Quem fala com o ERP é o `mc_api.MCApi`; quem
mostra na tela é o `pagamentos_frame.py`. Assim isto aqui roda em teste sem
navegador nenhum.

O QUE CADA LINHA RESPONDE
-------------------------
"Como eu pago isto, e o documento anexado confere com o lançamento?"

  Tipo de Pgto   Boleto ou Pix — e o BOLETO GANHA sempre que houver um anexado
  Dados do Pgto  a linha digitável ou a chave Pix
  Valor
  Descrição      CENTRO DE CUSTO  NF <nº>  OC <nº>
                 mão de obra:     CENTRO DE CUSTO  C <contrato>  M <medição>
                 água/luz:        CENTRO DE CUSTO  <descrição com UC e mês>
  Favorecido
  Status         APTO / ATENÇÃO / JÁ PAGO
  Conferência    o cruzamento documento × lançamento
  Obs            avisos, e a observação escrita no próprio lançamento

CRUZAMENTO: "DIVERGE" NÃO É "NÃO VERIFIQUEI"
--------------------------------------------
`DIVERGE` = o documento CONTRADIZ o lançamento (nº de NF diferente, CNPJ do
emitente diferente da chave Pix, UC de outra unidade). Vira ATENÇÃO.
`?` = não deu para verificar (anexo é foto, PDF sem texto, campo em branco).
Não vira alarme. Alarme falso ensina o usuário a ignorar alarme.
"""
from __future__ import annotations

import io
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import NamedTuple

_AQUI = Path(__file__).resolve().parent
for _p in (_AQUI, _AQUI.parent):         # a pasta da aba e a raiz do projeto
    if str(_p) not in sys.path:          # (rodando este módulo isoladamente)
        sys.path.insert(0, str(_p))

import util                              # noqa: E402  utilitário compartilhado

# Irmãos, na mesma pasta. O sufixo `_pagamento` não é enfeite: nome de módulo
# é global no sys.path, `aportes/regras.py` já existe e `pagamentos_dia` entra
# ANTES de `aportes` no caminho de import — um `regras.py` aqui seria
# importado no lugar dele e quebraria a aba Aportes.
import ocr_boleto                        # noqa: E402
import regras_pagamento as regras        # noqa: E402


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except ImportError:                                   # pragma: no cover
    pdfplumber = None


# --------------------------------------------------------------------------
# Texto
# --------------------------------------------------------------------------
sem_acento = util.sem_acento


def chave(s: str | None) -> str:
    """Forma comparável usada AQUI dentro.

    Difere do `util.norm_espaco` de propósito: aqui a comparação é entre
    textos livres da API (método de pagamento, nome de conta como veio), e
    colapsar espaços internos deixaria dois valores diferentes iguais. O que
    importa é ser a mesma função dos dois lados da comparação."""
    return util.sem_acento(s).casefold().strip()


def brl(v) -> str:
    try:
        return ("R$ " + f"{float(v):,.2f}").replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return str(v)


def para_data(valor) -> date | None:
    """Aceita ISO, ISO com hora, dd/mm/aaaa e epoch em ms."""
    if valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        try:
            return datetime.fromtimestamp(valor / 1000).date()
        except (OverflowError, OSError, ValueError):
            return None
    texto = str(valor).strip()
    for padrao in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto[:10], padrao).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------
# Contas e período
# --------------------------------------------------------------------------
CONTAS_IGNORAR = re.compile(r"APENAS\s+LANÇAMENTO|APENAS\s+AJUSTE|ERRADA", re.I)

#: `dateField` do ERP -> campo correspondente no item devolvido.
_DATEFIELD_CAMPO = {"PLANNED": "plannedDate", "PAYMENT": "dateOfPayment",
                    "REFERENCE": "referenceDate", "DUE": "dueDate"}
_CAMPOS_DATA = ("plannedDate", "dueDate", "dateOfPayment", "referenceDate")


def data_do_item(item: dict, preferido: str = "plannedDate") -> date | None:
    for campo in [preferido, *_CAMPOS_DATA]:
        if campo in item:
            d = para_data(item[campo])
            if d:
                return d
    return None


def filtrar_periodo(lancamentos, inicio: date, fim: date,
                    campo: str = "plannedDate", log=print) -> list[dict]:
    """Rede de segurança: confere o período item a item, no cliente.

    Se a API ignorar o filtro de data, o relatório não pode sair errado em
    silêncio — os itens de fora são descartados e o programa avisa quantos.
    """
    dentro, fora, sem_data = [], 0, 0
    for item in lancamentos:
        d = data_do_item(item, campo)
        if d is None:
            sem_data += 1
            dentro.append(item)
        elif inicio <= d <= fim:
            dentro.append(item)
        else:
            fora += 1
    if fora:
        log(f"  [!] {fora} lançamento(s) fora de {inicio:%d/%m} a {fim:%d/%m} "
            f"descartados (a API ignorou o filtro de '{campo}').")
    if sem_data:
        log(f"  {sem_data} lançamento(s) sem data reconhecida; mantidos.")
    return dentro


def separar_pagos(lancamentos) -> tuple[list, list]:
    pagos = [i for i in lancamentos if i.get("paid")]
    return [i for i in lancamentos if not i.get("paid")], pagos


def nome_da_conta(item: dict) -> str:
    return ((item.get("tradePayableAccount") or {}).get("name") or "SEM CONTA").strip()


def conta_entra(conta: str, incluir=(), excluir=()) -> bool:
    """Casa por PEDAÇO do nome, sem acento e sem caixa. Excluir vence incluir."""
    alvo = chave(conta)
    if CONTAS_IGNORAR.search(conta):
        return False
    if any(e in alvo for e in excluir):
        return False
    return not incluir or any(i in alvo for i in incluir)


def resumo_por_conta(lancamentos) -> list[tuple]:
    """[(conta, a_pagar, total, ja_pagos, ignorada)] ordenado por valor."""
    contas = defaultdict(lambda: [0, 0.0, 0])
    for it in lancamentos:
        c = contas[nome_da_conta(it)]
        if it.get("paid"):
            c[2] += 1
        else:
            c[0] += 1
            c[1] += valor_do_item(it)
    return sorted(((n, q, t, p, bool(CONTAS_IGNORAR.search(n)))
                   for n, (q, t, p) in contas.items()), key=lambda r: -r[2])


# --------------------------------------------------------------------------
# Valor e tipo de pagamento
# --------------------------------------------------------------------------
def valor_do_item(item: dict) -> float:
    """`remainingValue` é o que FALTA pagar: vem 0.0 em título já quitado, e
    aí o valor real está em sumOfPaidValues."""
    for campo in ("remainingValue", "sumOfPaidValues", "sumOfPaids", "plannedValue"):
        try:
            v = float(item.get(campo) or 0)
        except (TypeError, ValueError):
            continue
        if v:
            return v
    return 0.0


_E_BOLETO = re.compile(r"boleto|blt|cobran|febraban|fatura", re.I)
_NAO_E_BOLETO = re.compile(
    r"nota fiscal|nfe|danfe|recibo|comprovante|contrato|medi[çc][ãa]o|qr\s*code|pagar\s*para",
    re.I)


def _rotulo(f: dict) -> str:
    return f"{f.get('filename') or ''} {f.get('tagName') or ''}"


def tem_boleto(files) -> bool:
    return any(_E_BOLETO.search(_rotulo(f)) and not _NAO_E_BOLETO.search(_rotulo(f))
               for f in (files or ()))


def tipo_de_pagamento(item: dict, files=()) -> str:
    """Boleto ou Pix — e o BOLETO GANHA sempre que houver um anexado.

    `tradePayablePaymentMethod` não é confiável: um título com anexo
    `[Boleto] boleto oc 5909` vinha marcado como 'Pix' só porque o fornecedor
    tem chave no cadastro. Pagar por pix um título que veio com boleto é
    exatamente o erro que esta conferência existe para evitar.
    """
    if tem_boleto(files):
        return "Boleto"
    metodo = chave(item.get("tradePayablePaymentMethod"))
    if "pix" in metodo:
        return "Pix"
    if "boleto" in metodo:
        return "Boleto"
    if re.search(r"pix|chave", item.get("paidToBankAccount") or "", re.I):
        return "Pix"
    return (item.get("tradePayablePaymentMethod") or "Boleto").strip() or "Boleto"


def eh_pdf(f: dict) -> bool:
    # O ERP devolve extension COM ponto (".pdf"); comparar com "pdf" nunca bate.
    return ((f.get("extension") or "").strip().lstrip(".").lower() == "pdf"
            or (f.get("filename") or "").lower().endswith(".pdf"))


def escolher_pdf_do_boleto(files) -> dict | None:
    """O PDF do boleto. Nunca devolve a NF, mas não exige a palavra 'boleto':
    metade das faturas da concessionária vem com tagName nulo e nome só de
    número, e ficavam de fora."""
    explicitos = [f for f in files
                  if _E_BOLETO.search(_rotulo(f)) and not _NAO_E_BOLETO.search(_rotulo(f))]
    for f in explicitos:
        if eh_pdf(f):
            return f
    neutros = [f for f in files if eh_pdf(f) and not _NAO_E_BOLETO.search(_rotulo(f))]
    if len(neutros) == 1:
        return neutros[0]
    return explicitos[0] if explicitos else None


# --------------------------------------------------------------------------
# Chave Pix e linha digitável
# --------------------------------------------------------------------------
_PADROES_PIX = (
    re.compile(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}"),                       # CNPJ
    re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b"),                          # CPF
    re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+"),                                  # e-mail
    re.compile(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", re.I),
    re.compile(r"(?:\+?55\s*)?\(?\d{2}\)?\s*9?\d{4}[- ]?\d{4}"),              # celular
)
_PREFIXO_PIX = re.compile(
    r"^\s*(chave|pix|cnpj|cpf|celular|telefone|e-?mail|aleat\w*)\b[\s:.\-]*", re.I)


def parece_chave_pix(s: str) -> bool:
    """Chave Pix tem dígito ou '@'. 'VER COMENTÁRIO DA SOLICITAÇÃO' não é chave."""
    return bool(s) and bool(re.search(r"[\d@]", s))


def chave_pix_por_padrao(texto: str) -> str:
    """A primeira coisa com CARA de chave Pix no texto, ou "".

    Separada de `extrair_chave_pix` porque esta admite não achar nada. A
    outra, quando não acha padrão, devolve o texto limpo — útil para o campo
    do cadastro, perigoso para varrer um documento inteiro.
    """
    for padrao in _PADROES_PIX:
        m = padrao.search(texto or "")
        if m:
            return m.group(0).strip()
    return ""


def extrair_chave_pix(texto: str) -> str:
    """Devolve só a chave, sem o 'PIX CNPJ:' na frente."""
    achado = chave_pix_por_padrao(texto)
    if achado:
        return achado
    limpo = texto or ""
    for _ in range(4):                       # 'CHAVE PIX CELULAR :' empilha prefixos
        novo = _PREFIXO_PIX.sub("", limpo)
        if novo == limpo:
            break
        limpo = novo
    return limpo.strip()


#: Pix copia-e-cola (EMV): começa em 000201 e termina no CRC `6304XXXX`.
_PIX_COPIA_COLA = re.compile(r"00020[01][0-9A-Za-z._@+\-/*:]{20,500}?6304[0-9A-F]{4}")


def chave_pix_do_comentario(comentario: str) -> str:
    """A observação do lançamento às vezes É a forma de pagar, não um recado."""
    if not comentario:
        return ""
    m = _PIX_COPIA_COLA.search(re.sub(r"\s+", "", comentario))
    if m:
        return m.group(0)
    if re.search(r"\b(pix|chave)\b", comentario, re.I):
        achado = extrair_chave_pix(comentario)
        if achado and achado != comentario.strip():
            return achado
    return ""


_LINHAS_DIGITAVEIS = (
    re.compile(r"\d{5}\.?\d{5}\s+\d{5}\.?\d{6}\s+\d{5}\.?\d{6}\s+\d\s+\d{14}"),
    re.compile(r"\d{11,12}[- ]\d\s+\d{11,12}[- ]\d\s+\d{11,12}[- ]\d\s+\d{11,12}[- ]\d"),
    re.compile(r"\b\d{47,48}\b"),
)


def extrair_linha_digitavel(texto: str) -> str:
    for padrao in _LINHAS_DIGITAVEIS:
        m = padrao.search(texto or "")
        if m:
            return re.sub(r"\s+", " ", m.group(0)).strip()
    return ""


def texto_de_pdf(dados: bytes) -> str:
    if not (pdfplumber and dados):
        return ""
    try:
        with pdfplumber.open(io.BytesIO(dados)) as pdf:
            return "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    except Exception:
        return ""


# --------------------------------------------------------------------------
# Reembolso ("PAGAR PARA <nome>")
# --------------------------------------------------------------------------
def classificar_anexos(files) -> str:
    if not files:
        return "SEM_ANEXO"
    s = " | ".join(chave(_rotulo(f)) for f in files)
    if "autorizado" in s:
        return "AUTORIZADO"
    if re.search(r"pagar\s*_?\s*para", s):      # antes de NF: reembolso manda
        return "PAGAR_PARA"
    if re.search(r"nfe|danfe|nota fiscal", s):
        return "NF"
    if re.search(r"boleto|blt", s):
        return "BOLETO"
    if "recibo" in s:
        return "RECIBO"
    return "OUTRO"


def nome_do_reembolso(files) -> str:
    for f in files:
        m = re.search(r"pagar\s*_?\s*para\s*[-:_ ]*(.+)", chave(f.get("filename")))
        if m:
            return re.sub(r"\.(pdf|jpe?g|png|docx?)$", "", m.group(1)).strip()
    return ""


def pix_do_reembolso(files, item: dict, mapa: dict) -> str:
    """A chave Pix da PESSOA do aviso — nunca a do cadastro do fornecedor."""
    alvo = nome_do_reembolso(files) + " " + chave(item.get("documentNumber"))
    for nome, pix in sorted(mapa.items(), key=lambda kv: -len(kv[0])):
        if chave(nome) in alvo:
            return pix
    return ""


_PAGAR_PARA = re.compile(r"pagar\s*_?\s*para", re.I)


def chave_pix_do_aviso(files, textos: dict) -> str:
    """O número escrito DENTRO do aviso "PAGAR PARA", logo abaixo do nome.

    Quem monta o aviso já escreve ali o CPF ou o celular de quem recebe. O
    relatório só sabia consultar o `pix_reembolso.json`, que depende de
    alguém ter cadastrado a pessoa antes — então reembolso de gente nova
    chegava sempre como "chave não cadastrada; abrir o aviso".

    A busca fica na JANELA logo após o "pagar para": documento de reembolso
    costuma trazer também o CNPJ da empresa e o valor, e varrer o texto
    inteiro pegaria o primeiro número parecido, não o certo.
    """
    for f in files or ():
        if not _PAGAR_PARA.search(_rotulo(f)):
            continue
        texto = textos.get(f.get("downloadUrl") or "") or ""
        m = _PAGAR_PARA.search(texto)
        janela = texto[m.end():m.end() + 300] if m else ""
        achado = chave_pix_por_padrao(janela)
        if achado:
            return achado
    return ""


def mesma_chave(a: str, b: str) -> bool:
    """Duas grafias da mesma chave? '111.222.333-44' e 'Fulano 11122233344'."""
    da, db = re.sub(r"\D", "", a or ""), re.sub(r"\D", "", b or "")
    if da and db:
        return da in db or db in da
    return chave(a) == chave(b)


# --------------------------------------------------------------------------
# Descrição
# --------------------------------------------------------------------------
_OC_NO_NOME = re.compile(
    r"\b(?:oc|ordem\s+de\s+compra)\s*[:\-–]?\s*n?[ºo°]?\s*(\d{2,7})\b", re.I)
_DOC_NO_NOME = re.compile(r"\bN[ºo°F]\s*[:\-]?\s*(\d{2,10})\b", re.I)

_UTILIDADES = re.compile(
    r"sanesc|saneago|equatorial|enel|celg|cemig|copasa|caesb|energisa|"
    r"neoenergia|light|sabesp|corsan|compesa|cagece|embasa", re.I)
_UC = re.compile(r"\bUC\s*[:\-]?\s*(\d{6,15})\b", re.I)


def eh_utilidade(item: dict) -> bool:
    return bool(_UTILIDADES.search(sem_acento(item.get("paidTo") or "")))


def unidade_consumidora(item: dict) -> str:
    m = _UC.search(item.get("description") or "")
    return m.group(1) if m else ""


def comentario_do(overview) -> str:
    return ((overview or {}).get("comment") or "").strip()


def achar_oc(item: dict, files, comentario: str = "", overview=None) -> str:
    """A fonte boa é `overview.purchaseOrder.number`; o resto é rede de
    segurança para quando o detalhe não carregar. Não exige
    `hasPurchaseOrder`: a flag às vezes vem falsa em título que tem OC."""
    numero = ((overview or {}).get("purchaseOrder") or {}).get("number")
    if numero:
        return str(numero).strip()
    for texto in [f.get("filename") or "" for f in files] + \
                 [item.get("description") or "", comentario]:
        m = _OC_NO_NOME.search(texto)
        if m:
            return m.group(1)
    return ""


def achar_doc(item: dict, files, overview=None) -> str:
    doc = (item.get("documentNumber") or "").strip()
    if doc and not re.search(r"REEMBOLSO", doc, re.I):
        return doc
    for f in files:
        m = _DOC_NO_NOME.search(f.get("filename") or "")
        if m:
            return m.group(1)
    return str((overview or {}).get("documentNumber") or "").strip()


def centro_de_custo(item: dict) -> str:
    """O centro de custo, uma vez só.

    Lançamento rateado repete o MESMO imóvel em cada parcela do rateio, e a
    descrição saía com o endereço duplicado ("TB 21 QD 51 LT 38 | TB 21 QD 51
    LT 38 NF 61106"). Rateio entre imóveis DIFERENTES continua mostrando os
    dois — é informação, não repetição."""
    nomes = [(c.get("workName") or "").strip()
             for c in (item.get("costCentreDetails") or []) if c]
    return " | ".join(dict.fromkeys(n for n in nomes if n))


def monta_descricao(item: dict, files, comentario: str = "", overview=None) -> str:
    cc = centro_de_custo(item)
    doc = achar_doc(item, files, overview)
    oc = achar_oc(item, files, comentario, overview)

    # "OC5928" no campo Número do documento não é número de nota. Sem isto a
    # planilha anunciava o mesmo número duas vezes ("NF 5928 OC 5928") e a
    # conferência de NF procurava uma nota que não existe.
    if doc and regras.documento_e_a_oc(doc, oc):
        oc = oc or regras.oc_no_documento(doc) or re.sub(r"\D", "", doc)
        doc = ""

    # Água/energia: o que identifica é a descrição (UC, mês, casa). O "número
    # da NF" ali é o número da fatura e não ajuda ninguém a conferir.
    if eh_utilidade(item):
        partes = [p for p in (cc, (item.get("description") or "").strip()) if p]
        if oc:
            partes.append(f"OC {oc}")
        return " ".join(partes).strip()

    partes = [cc] if cc else []
    if doc:
        partes.append(f"NF {doc}")
    if oc:
        partes.append(f"OC {oc}")
    if not doc and not oc:
        m = re.search(r"-\s*(\d+)\s*-\s*Medi[çc][ãa]o:\s*(\d+)",
                      item.get("description") or "")
        if m:
            partes += [f"C {m.group(1)}", f"M {m.group(2)}"]
        elif item.get("description"):
            # 40 caracteres cortavam exatamente onde mora o que distingue as
            # linhas ("... - CASA 1/2/3"), deixando-as idênticas na planilha.
            partes.append(item["description"].strip()[:110])
    return " ".join(partes).strip()


# --------------------------------------------------------------------------
# Cruzamento documento × lançamento
# --------------------------------------------------------------------------
_CHAVE_NFE = re.compile(r"(?<!\d)(\d{44})(?!\d)")
_RUIDO_RAZAO = {
    "ltda", "me", "epp", "eireli", "cia", "com", "comercio", "industria",
    "servicos", "distribuidora", "materiais", "construcao", "empreendimentos",
    "engenharia", "produtos", "grupo", "casa",
}
_RUIDO_ENDERECO = {"casa", "obra", "quadra", "lote", "bloco", "pos", "area",
                   "terreno", "etapa", "fase", "loteamento", "residencial"}


def dados_da_chave_nfe(chave44: str) -> dict:
    """Desmonta a chave de acesso da NF-e (44 dígitos).

    cUF(2) AAMM(4) CNPJ(14) mod(2) série(3) nNF(9) tpEmis(1) cNF(8) cDV(1).
    Vem de graça no nome do anexo: dá para conferir número e emitente sem
    baixar PDF nenhum.
    """
    return {"uf": chave44[0:2], "aamm": chave44[2:6], "cnpj": chave44[6:20],
            "modelo": chave44[20:22], "serie": chave44[22:25],
            "numero": str(int(chave44[25:34]))}


def _cnpj_fmt(c: str) -> str:
    return f"{c[0:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}" if len(c) == 14 else c


def _valor_nos_textos(valor: float, textos) -> bool:
    if not valor:
        return False
    br = f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    junto = " ".join(textos)
    return any(a in junto for a in {br, br.replace(".", ""),
                                    f"{valor:.2f}".replace(".", ",")})


def _enderecos(item: dict) -> tuple[list[str], list[str]]:
    """Duas formas de identificar o imóvel, porque as duas pontas escrevem
    diferente: estruturado (`QD 18`, `LT 8`) como o ERP escreve, e o nome do
    logradouro como a conta da concessionária escreve."""
    cc = centro_de_custo(item)
    fonte = sem_acento(f"{cc} {item.get('description') or ''}")
    estruturados = re.findall(r"\b(?:QD|LT|CASA|CS)\s*\.?\s*\d{1,4}\b", fonte, re.I)
    ruas = [p for p in re.findall(r"[A-Za-z]{4,}", sem_acento(cc))
            if p.lower() not in _RUIDO_ENDERECO]
    return estruturados, ruas


def _normaliza_para_busca(textos) -> str:
    return re.sub(r"\s+", " ", sem_acento(" ".join(textos)).upper())


def _conferir_endereco(item: dict, textos) -> str:
    """"endereço ✓" / "endereço ?" para QUALQUER linha com QD/LT.

    Há fornecedor que anexa o pedido, e no pedido está o endereço da obra —
    daí a pergunta "o boleto é mesmo desta casa?". Ela vale para todo
    fornecedor que anexe pedido ou orçamento, então a checagem é geral.

    NUNCA vira DIVERGE: o anexo escreve "QUADRA 40" onde o ERP escreveu
    "QD 40", e alarme falso ensina o usuário a ignorar alarme. Aqui só se
    informa o que foi possível confirmar.
    """
    estruturados = dict.fromkeys(re.sub(r"\s+", " ", p.upper())
                                 for p in _enderecos(item)[0])
    if not estruturados:
        return ""
    alvo = _normaliza_para_busca(textos)
    faltando = [p for p in estruturados if p not in alvo]
    if not faltando:
        return f"endereço ✓ ({', '.join(estruturados)[:40]})"
    return f"endereço ? ({', '.join(faltando)[:40]} não aparece)"


def _conferir_utilidade(item, files, textos) -> tuple[list[str], bool]:
    """Conta de água/luz: o que identifica é a UC e o endereço, não a 'NF'.
    A UC aparece no NOME do anexo (`000451784501210` para `UC 451784501210`)."""
    partes, divergiu = [], False
    uc = unidade_consumidora(item)
    if not uc:
        partes.append("UC ? (não informada na descrição)")
    else:
        onde = " ".join([f.get("filename") or "" for f in files] + list(textos))
        if uc in re.sub(r"\D", "", onde) or uc in onde:
            partes.append(f"UC {uc} ✓")
        else:
            partes.append(f"UC DIVERGE: {uc} não aparece no anexo")
            divergiu = True

    estruturados, ruas = _enderecos(item)
    if not (estruturados or ruas):
        partes.append("endereço ? (lançamento não diz)")
    elif not textos:
        partes.append("endereço ? (anexo sem texto)")
    else:
        alvo = re.sub(r"\s+", " ", sem_acento(" ".join(textos)).upper())
        achados = [p for p in estruturados if re.sub(r"\s+", " ", p.upper()) in alvo]
        achados += [r for r in ruas if r.upper() in alvo]
        if achados:
            partes.append(f"endereço ✓ ({', '.join(dict.fromkeys(achados))[:40]})")
        else:
            partes.append(f"endereço ? ({', '.join((estruturados + ruas)[:3])} não aparece)")
    return partes, divergiu


def conferir_documento(item: dict, files, textos, overview=None) -> tuple[str, bool]:
    """Cruza o anexo com o lançamento. Devolve (resumo, tem_divergência)."""
    partes, divergiu = [], False
    textos = [t for t in textos if t]

    if eh_utilidade(item):
        partes, divergiu = _conferir_utilidade(item, files, textos)
        if textos:
            partes.append("valor ✓" if _valor_nos_textos(valor_do_item(item), textos)
                          else "valor ?")
        return " · ".join(partes), divergiu

    # NF: número e emitente, direto da chave de acesso do nome do anexo.
    chaves = [m.group(1) for f in files
              for m in _CHAVE_NFE.finditer(f.get("filename") or "")]
    doc = (item.get("documentNumber") or "").strip()

    # "OC5928" no campo do documento não é número de nota: é a ordem de
    # compra, e tem de ser a MESMA do lançamento. Dois números diferentes ali
    # significam que um dos dois está errado, e alguém precisa olhar.
    oc_real = achar_oc(item, files, "", overview)
    if doc and regras.documento_e_a_oc(doc, oc_real):
        oc_escrita = regras.oc_no_documento(doc) or re.sub(r"\D", "", doc)
        if oc_real and oc_escrita.lstrip("0") != str(oc_real).lstrip("0"):
            partes.append(f"OC DIVERGE: documento {doc} × ordem {oc_real}")
            divergiu = True
        else:
            partes.append(f"OC {oc_escrita} ✓")
        doc = ""                         # não procurar nota fiscal que não existe

    if chaves:
        nfe = dados_da_chave_nfe(chaves[0])
        if doc.isdigit():
            if nfe["numero"] == str(int(doc)):
                partes.append(f"NF {nfe['numero']} ✓")
            else:
                partes.append(f"NF DIVERGE: anexo {nfe['numero']} × lançamento {doc}")
                divergiu = True
        else:
            partes.append(f"NF {nfe['numero']} (lançamento sem nº)")

        achado = re.search(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}",
                           item.get("paidToBankAccount") or "")
        cnpj_pix = re.sub(r"\D", "", achado.group(0)) if achado else ""
        if cnpj_pix:
            if cnpj_pix == nfe["cnpj"]:
                partes.append("CNPJ ✓")
            else:
                partes.append(f"CNPJ DIVERGE: NF {_cnpj_fmt(nfe['cnpj'])} × "
                              f"pix {_cnpj_fmt(cnpj_pix)}")
                divergiu = True
    elif doc:
        partes.append("NF ? (sem chave no anexo)")

    if not textos:
        partes.append("valor/fornecedor ? (anexo sem texto)")
        return " · ".join(partes), divergiu

    partes.append("valor ✓" if _valor_nos_textos(valor_do_item(item), textos) else "valor ?")

    alvo = chave(" ".join(textos))
    esperados = [p for p in re.findall(r"[a-z0-9]+", chave(item.get("paidTo") or ""))
                 if len(p) >= 4 and p not in _RUIDO_RAZAO]
    if not esperados:
        partes.append("fornecedor ?")
    elif any(t in alvo for t in esperados):
        partes.append("fornecedor ✓")
    else:
        partes.append(f"fornecedor ? ({(item.get('paidTo') or '')[:20]} não aparece)")

    endereco = _conferir_endereco(item, textos)
    if endereco:
        partes.append(endereco)
    return " · ".join(partes), divergiu


# --------------------------------------------------------------------------
# Montagem das linhas
# --------------------------------------------------------------------------
class Resultado(NamedTuple):
    """O que sai da regra de negócio: o que entra e o que ficou de fora.

    Os omitidos viajam junto de propósito. Enquanto "não entrou" era só um
    `continue`, a única forma de descobrir que uma regra tinha errado era
    sentir falta de um pagamento — o que só acontece depois do vencimento.
    """
    contas: dict          # {conta: [linha, ...]}
    omitidos: list        # [{conta, tipo, valor, descricao, favorecido, motivo}]


def montar_registros(lancamentos, anexos: dict, overviews: dict, textos: dict,
                     incluir=(), excluir=(), pix_reembolso=None,
                     urls_ocr=(), regras_fornecedor=None,
                     ids_nao_confirmados=()) -> Resultado:
    """Transforma lançamentos do ERP em linhas de planilha.

    `anexos`             {tradePayableId: [anexo]}
    `overviews`          {installmentId: overview}
    `textos`             {downloadUrl: texto do PDF} — vazio pula o cruzamento
    `urls_ocr`           quais desses textos vieram de OCR (leitura duvidosa,
                         que só é aceita depois de conferir DV e valor)
    `regras_fornecedor`  cadastro de `regras_pagamento.carregar_fornecedores`
    `ids_nao_confirmados` lançamentos que a pessoa desmarcou na janela de
                         confirmação — saem com motivo, não em silêncio
    """
    pix_reembolso = pix_reembolso or {}
    regras_forn = regras_fornecedor or {}
    ids_nao_confirmados = {str(i) for i in ids_nao_confirmados}
    registros, omitidos = defaultdict(list), []

    for item in lancamentos:
        conta = nome_da_conta(item)
        if not conta_entra(conta, incluir, excluir):
            continue

        files = anexos.get(str(item.get("tradePayableId"))) or []
        overview = overviews.get(str(item.get("id"))) or {}
        coment = comentario_do(overview)
        cls = classificar_anexos(files)
        tipo = tipo_de_pagamento(item, files)
        pago_para = (item.get("paidToBankAccount") or "").strip()
        valor = valor_do_item(item)
        favorecido = (item.get("paidTo") or "").strip()

        do_item = [textos.get(f.get("downloadUrl") or "") for f in files]
        do_item = [t for t in do_item if t] + ([coment] if coment else [])
        conferencia, divergiu = (conferir_documento(item, files, do_item, overview)
                                 if textos else ("(não cruzado)", False))

        # A compra está documentada? É o que decide se um título sem boleto
        # anexado pode ser pago pela chave do cadastro (abaixo) ou se é ruído.
        nf = achar_doc(item, files, overview)
        oc = achar_oc(item, files, coment, overview)
        tem_nf_ou_oc = bool(oc or (nf and not regras.documento_e_a_oc(nf, oc)))

        avisos, obs, chave_divergente = [], "", False
        #: Há um documento anexado do qual dá para tirar a forma de pagar à
        #: mão? Separa "o boleto veio como foto e o OCR não fechou" (fica na
        #: planilha, para alguém abrir e digitar) de "não veio boleto nenhum"
        #: (não fica: não há o que digitar, e a linha só custa conferência).
        tem_documento = False

        if cls == "PAGAR_PARA":
            tipo = "Pix"
            tem_documento = True
            do_aviso = chave_pix_do_aviso(files, textos)
            do_mapa = pix_do_reembolso(files, item, pix_reembolso)
            if do_aviso and do_mapa and not mesma_chave(do_aviso, do_mapa):
                dados, chave_divergente = do_aviso, True
                obs = (f"Reembolso — a chave do AVISO ({do_aviso}) difere da "
                       f"cadastrada ({do_mapa}); confirmar antes de pagar")
            elif do_aviso:
                dados = do_aviso
                obs = "Reembolso — chave lida do próprio aviso, NÃO o pix do cadastro"
            elif do_mapa:
                dados = do_mapa
                obs = "Reembolso — pagar a chave do aviso, NÃO o pix do cadastro"
            else:
                dados = ""
                obs = (f"Reembolso para '{nome_do_reembolso(files) or '?'}' — chave não "
                       "cadastrada; abrir o aviso")
        elif tipo == "Pix":
            dados = extrair_chave_pix(pago_para) if pago_para else ""
            if not parece_chave_pix(dados):
                # O cadastro às vezes traz um recado ("VER COMENTÁRIO DA
                # SOLICITAÇÃO") no lugar da chave — a chave está na observação.
                do_comentario = chave_pix_do_comentario(coment)
                if do_comentario:
                    if dados:
                        avisos.append(f"Cadastro dizia: {dados}")
                    dados = do_comentario
                    avisos.append("Chave/copia-e-cola veio da observação do lançamento.")
                else:
                    dados, obs = "", "Pix sem chave no cadastro — buscar no ERP"
        else:
            pdf = escolher_pdf_do_boleto(files)
            url_pdf = (pdf or {}).get("downloadUrl") or ""
            texto_pdf = textos.get(url_pdf) or ""
            tem_documento = bool(pdf)
            if url_pdf in urls_ocr:
                # Texto de OCR NUNCA passa pelo extrator solto: ali um "8"
                # lido como "B" viraria linha digitável de mentira, com cara
                # de verdade. Quem lê OCR é quem também confere DV e valor.
                dados = ocr_boleto.achar_linha_digitavel(texto_pdf, valor)
                if dados:
                    avisos.append("Linha digitável lida por OCR — dígito verificador "
                                  "e valor conferem com o lançamento.")
            else:
                dados = extrair_linha_digitavel(texto_pdf)
            if not dados:
                do_cadastro = extrair_chave_pix(pago_para) if pago_para else ""
                if pdf:
                    obs = ("Boleto em imagem e o OCR não fechou — preencher manual"
                           if url_pdf in urls_ocr else
                           "Boleto em imagem — preencher manual")
                elif parece_chave_pix(do_cadastro) and tem_nf_ou_oc:
                    # Sem boleto anexado, a regra "boleto ganha de Pix" não
                    # tem premissa: não há boleto para ganhar. Havendo NF ou
                    # OC, a compra está documentada e o Pix do cadastro é a
                    # forma de pagar. Sem NF nem OC, a linha não entra.
                    tipo, dados = "Pix", do_cadastro
                    obs = "Sem boleto anexado — pagar pela chave Pix do cadastro"
                else:
                    obs = "Nenhum anexo de boleto — conferir no ERP"

        # Depois de resolver a forma de pagar, e não antes: quando a linha
        # acabou virando Pix por falta de boleto, mandar "pagar o boleto"
        # seria mandar pagar um documento que não existe.
        if tipo == "Boleto" and re.search(r"pix|chave", pago_para, re.I) and cls != "PAGAR_PARA":
            avisos.append(f"Cadastro tem Pix ({extrair_chave_pix(pago_para)}), "
                          "mas o título veio com BOLETO — pagar o boleto.")

        # O código de barras é conferido por dígito verificador e carrega o
        # valor: quando ele discorda do lançamento, quem erra é o lançamento.
        # Vale para a linha que veio do PDF com texto, não só para a do OCR —
        # a do OCR já nasceu conferida contra o valor (`linha_confiavel`).
        valor_documento = (ocr_boleto.valor_da_linha(dados)
                           if dados and ocr_boleto.valida(dados) else None)
        valor_diverge = regras.documento_contradiz_valor(valor, valor_documento)
        if valor_diverge:
            avisos.append(f"Boleto diz {brl(valor_documento)} e o lançamento "
                          f"diz {brl(valor)} — conferir ANTES de pagar.")

        # Tipo da chave Pix: o ERP não tem o campo, e sem ele CPF e celular
        # são os mesmos onze dígitos. Só avisa quando HÁ chave e o tipo não
        # dá para saber — nas outras não há o que perguntar.
        if tipo == "Pix" and regras.chave_pix_ambigua(pago_para or dados, dados):
            avisos.append("Chave Pix sem tipo declarado — confirmar se é CPF, "
                          "celular ou aleatória.")

        if coment:
            avisos.append(f"Observação do lançamento: {coment[:220]}")
        obs = " · ".join(filter(None, [obs] + avisos))

        descricao = monta_descricao(item, files, coment, overview)

        # Já pago é informação, não pagamento: as regras de omissão não valem
        # ali. Uma linha "JÁ PAGO" sem forma de pagar é o normal, não um erro.
        motivo = ""
        if not item.get("paid"):
            if str(item.get("id")) in ids_nao_confirmados:
                motivo = regras.MOTIVO_NAO_CONFIRMADO
            else:
                motivo = regras.motivo_omissao(valor, favorecido, dados,
                                               tem_documento, regras_forn,
                                               valor_documento=valor_documento)
        if motivo:
            omitidos.append({"conta": conta, "tipo": tipo, "valor": valor,
                             "descricao": descricao, "favorecido": favorecido,
                             "motivo": motivo})
            continue

        status = {"SEM_ANEXO": "ATENÇÃO — sem anexo",
                  "AUTORIZADO": "APTO (autorizado)",
                  "PAGAR_PARA": "APTO* (reembolso)"}.get(cls, "APTO")
        if not dados:
            status = "ATENÇÃO — sem dados de pgto"
        if divergiu:
            status = "ATENÇÃO — documento não bate"
        if chave_divergente:
            status = "ATENÇÃO — chave do reembolso divergente"
        # Por último entre os alarmes: pagar o valor errado é o pior deles.
        if valor_diverge:
            status = "ATENÇÃO — valor do boleto diverge"
        if item.get("paid"):
            # dateOfPayment as vezes vem em epoch (numero), e concatenar
            # numero com string levanta TypeError no meio do relatorio.
            _pago_em = para_data(item.get("dateOfPayment"))
            status = "JÁ PAGO em " + (f"{_pago_em:%d/%m/%Y}" if _pago_em
                                      else "?")

        registros[conta].append({
            "tipo": tipo, "dados": dados, "valor": valor,
            "descricao": descricao, "favorecido": favorecido,
            "status": status, "conferencia": conferencia, "obs": obs,
        })

    for regs in registros.values():
        regs.sort(key=lambda r: (r["tipo"], r["favorecido"]))
    omitidos.sort(key=lambda o: (o["conta"], o["motivo"], o["favorecido"]))
    return Resultado(dict(sorted(registros.items())), omitidos)


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------
HEADERS = ["Tipo de Pgto", "Dados do Pgto", "Valor", "Descrição", "Favorecido",
           "Status", "Conferência", "Obs"]
LARGURAS = [13, 40, 14, 52, 28, 24, 40, 42]
COL_VALOR = HEADERS.index("Valor") + 1

ABA_OMITIDOS = "NÃO ENTRARAM"
HEADERS_OMITIDOS = ["Conta", "Tipo de Pgto", "Valor", "Descrição", "Favorecido",
                    "Motivo de não ter entrado"]
LARGURAS_OMITIDOS = [34, 13, 14, 52, 28, 52]


def nomes_de_aba(contas) -> dict:
    """Nome de aba único: o Excel corta em 31 caracteres e não aceita repetido.
    'MORAIS ... - INTER' e '- SICOOB' cortam iguais e derrubavam a gravação."""
    usados, mapa = {ABA_OMITIDOS.casefold()}, {}
    for conta in contas:
        base = re.sub(r"[:\\/?*\[\]]", "-", conta or "SEM CONTA")[:31] or "SEM CONTA"
        nome, n = base, 2
        while nome.casefold() in usados:
            sufixo = f" ({n})"
            nome = base[:31 - len(sufixo)] + sufixo
            n += 1
        usados.add(nome.casefold())
        mapa[conta] = nome
    return mapa


def _formatar(ws, larguras, borda, primeira_linha=3):
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for linha in ws.iter_rows(min_row=primeira_linha):
        for cel in linha:
            cel.border = borda
            cel.alignment = Alignment(vertical="center", wrap_text=True)


def _aba_omitidos(wb, omitidos, cab, borda):
    """As linhas que as regras tiraram, com o motivo de cada uma.

    Existe para a regra poder ser conferida — e corrigida — por quem usa, em
    vez de o pagamento simplesmente não aparecer. Não soma no TOTAL de conta
    nenhuma: nada aqui é para pagar hoje.
    """
    ws = wb.create_sheet(ABA_OMITIDOS)
    ws.append([f"{len(omitidos)} lançamento(s) que as regras deixaram de fora "
               "— confira se alguma linha deveria estar sendo paga"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(HEADERS_OMITIDOS)
    for c in range(1, len(HEADERS_OMITIDOS) + 1):
        cel = ws.cell(row=3, column=c)
        cel.fill, cel.font = cab[0], cab[1]
        cel.alignment = Alignment(horizontal="center")
    for o in omitidos:
        ws.append([o["conta"], o["tipo"], o["valor"], o["descricao"],
                   o["favorecido"], o["motivo"]])
    _formatar(ws, LARGURAS_OMITIDOS, borda)
    for linha in ws.iter_rows(min_row=4, min_col=3, max_col=3):
        for cel in linha:
            if isinstance(cel.value, (int, float)):
                cel.number_format = 'R$ #,##0.00'
    ws.freeze_panes = "A4"


def gerar_excel(resultado, caminho: Path, log=print) -> Path:
    registros = getattr(resultado, "contas", resultado)
    omitidos = list(getattr(resultado, "omitidos", ()) or ())
    wb = Workbook()
    wb.remove(wb.active)
    fundo_cab = PatternFill("solid", fgColor="1F4E78")
    fonte_cab = Font(color="FFFFFF", bold=True)
    fundo_atencao = PatternFill("solid", fgColor="FCE4D6")
    fino = Side(style="thin", color="D9D9D9")
    borda = Border(left=fino, right=fino, top=fino, bottom=fino)
    abas = nomes_de_aba(registros.keys())

    for conta, regs in registros.items():
        ws = wb.create_sheet(abas[conta])
        ws.append([f"Conta: {conta}"])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])
        ws.append(HEADERS)
        for c in range(1, len(HEADERS) + 1):
            cel = ws.cell(row=3, column=c)
            cel.fill, cel.font = fundo_cab, fonte_cab
            cel.alignment = Alignment(horizontal="center")

        for r in regs:
            ws.append([r["tipo"], r["dados"], r["valor"], r["descricao"],
                       r["favorecido"], r["status"], r["conferencia"], r["obs"]])
            if r["status"].startswith("ATEN"):
                for c in range(1, len(HEADERS) + 1):
                    ws.cell(row=ws.max_row, column=c).fill = fundo_atencao

        ws.append([])
        total = ["TOTAL"] + [""] * (len(HEADERS) - 1)
        total[COL_VALOR - 1] = sum(r["valor"] for r in regs)
        ws.append(total)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=COL_VALOR).font = Font(bold=True)

        _formatar(ws, LARGURAS, borda)
        for linha in ws.iter_rows(min_row=4, min_col=COL_VALOR, max_col=COL_VALOR):
            for cel in linha:
                if isinstance(cel.value, (int, float)):
                    cel.number_format = 'R$ #,##0.00'
        ws.freeze_panes = "A4"

    if omitidos:
        _aba_omitidos(wb, omitidos, (fundo_cab, fonte_cab), borda)

    return _salvar(wb, caminho, log)


def _salvar(wb, caminho: Path, log=print) -> Path:
    """Grava o .xlsx; se estiver aberto no Excel, grava ao lado em vez de
    morrer. Perder uma coleta inteira por causa de um arquivo aberto na tela
    já aconteceu — não pode acontecer de novo."""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(caminho)
        return caminho
    except PermissionError:
        pass
    for n in range(2, 20):
        alternativo = caminho.with_name(f"{caminho.stem} ({n}){caminho.suffix}")
        try:
            wb.save(alternativo)
            log(f"  ({caminho.name} está aberto no Excel — salvei como {alternativo.name})")
            return alternativo
        except PermissionError:
            continue
    raise RuntimeError(f"Não consegui gravar {caminho.name}. Feche o Excel e tente de novo.")
