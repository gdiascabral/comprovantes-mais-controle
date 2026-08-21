"""Montagem da planilha final.

Escreve apenas a celula da data e as quatro colunas de dados (saldo, pagamento,
qtd sistema, qtd banco) nas linhas de conta. Todas as formulas do modelo — as
tres abas — sao preservadas, e o programa confere isso celula por celula depois
de salvar.
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .config import PlanilhaConfig
from .mapping import AccountMapping
from .models import RowFill
from .parsing import to_float


class WorkbookError(Exception):
    """Falha ao montar ou validar a planilha."""


@dataclass(frozen=True)
class BuildResult:
    path: Path
    linhas_escritas: int
    celulas_conferidas: int


def output_name(reference_date: date) -> str:
    """Nome idempotente do dia: "30 07 - completa.xlsx" (sobrescreve o do dia)."""
    return f"{reference_date:%d %m} - completa.xlsx"


def _comparavel(value: object) -> object:
    """Normaliza para comparar celulas relidas.

    O Excel guarda data como numero de serie e o openpyxl devolve `datetime`,
    mesmo quando gravamos um `date`.
    """
    if isinstance(value, datetime):
        return value.date()
    return value


def _sheet(wb, planilha: PlanilhaConfig) -> Worksheet:
    """A aba do painel, PELO NOME.

    Era `worksheets[0]` com a exigencia de aba unica. O modelo passou a ter tres
    abas (Painel, Movimentações, Regras) e a posicao de uma aba e coisa que se
    muda arrastando com o mouse — o nome nao.
    """
    if planilha.aba not in wb.sheetnames:
        raise WorkbookError(
            f"o modelo nao tem a aba {planilha.aba!r}; encontrei {wb.sheetnames}"
        )
    return wb[planilha.aba]


def check_labels(ws: Worksheet, mapping: AccountMapping) -> None:
    """Confere que o mapping.yaml casa com a coluna B do modelo.

    Critico: a ordem das linhas mudou uma vez (as quatro contas principais
    subiram para o topo) e o mapping guarda o numero da linha. Label fora de
    lugar poe o saldo de uma conta na linha de outra, sem erro nenhum no Excel.
    """
    divergencias = []
    for row in mapping.rows:
        no_modelo = ws[f"B{row.row}"].value
        if (no_modelo or "") != row.label:
            divergencias.append(f"  linha {row.row}: modelo={no_modelo!r} mapping={row.label!r}")

    if divergencias:
        raise WorkbookError(
            "label do mapping.yaml diverge da coluna B do modelo:\n" + "\n".join(divergencias)
        )


def build(
    modelo: str | Path,
    destino: str | Path,
    reference_date: date,
    fills: list[RowFill],
    mapping: AccountMapping,
    planilha: PlanilhaConfig,
) -> BuildResult:
    """Copia o modelo, preenche o dia e valida o resultado."""
    modelo, destino = Path(modelo), Path(destino)
    if not modelo.is_file():
        raise WorkbookError(f"modelo nao encontrado: {modelo}")

    destino.parent.mkdir(parents=True, exist_ok=True)
    # Copia antes de abrir para nunca correr risco de salvar sobre o original.
    shutil.copy2(modelo, destino)

    wb = openpyxl.load_workbook(destino)
    ws = _sheet(wb, planilha)
    check_labels(ws, mapping)

    # O modelo vem com formato americano (m/d/yyyy); forcamos DD/MM/YYYY para
    # ler certo em portugues.
    celula_data = ws[planilha.celula_data]
    celula_data.value = reference_date
    celula_data.number_format = planilha.formato_data

    escritas = 0
    for fill in fills:
        if fill.row not in planilha.linhas:
            raise WorkbookError(f"linha {fill.row} fora da faixa do painel")
        ws[f"{planilha.col_saldo}{fill.row}"] = to_float(fill.balance)
        ws[f"{planilha.col_pagamento}{fill.row}"] = to_float(fill.total)
        ws[f"{planilha.col_qtd_sistema}{fill.row}"] = fill.count
        ws[f"{planilha.col_qtd_banco}{fill.row}"] = fill.bank_count
        escritas += 1

    wb.save(destino)
    wb.close()

    conferidas = assert_untouched(modelo, destino, reference_date, fills, planilha)
    return BuildResult(path=destino, linhas_escritas=escritas, celulas_conferidas=conferidas)


def assert_untouched(
    modelo: str | Path,
    destino: str | Path,
    reference_date: date,
    fills: list[RowFill],
    planilha: PlanilhaConfig,
) -> int:
    """Reabre os dois arquivos e prova que so as celulas previstas mudaram.

    Devolve a quantidade de celulas conferidas. Qualquer formula perdida no
    round-trip do openpyxl aparece aqui como divergencia.

    Confere as TRES abas: o painel na area util, e as abas de apoio inteiras. A
    aba «Movimentações» carrega o rateio do aporte e a «Regras» diz quem paga —
    um VLOOKUP perdido ali sairia como zero na ordem de transferencia do dia.
    """
    esperado = _celulas_escritas(reference_date, fills, planilha)

    wb_orig = openpyxl.load_workbook(Path(modelo))
    wb_novo = openpyxl.load_workbook(Path(destino))
    try:
        if wb_orig.sheetnames != wb_novo.sheetnames:
            raise WorkbookError(
                f"abas mudaram: modelo={wb_orig.sheetnames} gerado={wb_novo.sheetnames}"
            )

        problemas: list[str] = []
        conferidas = 0

        ws_orig, ws_novo = _sheet(wb_orig, planilha), _sheet(wb_novo, planilha)
        (r1, r2), (c1, c2) = planilha.area_conferida
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                original = ws_orig.cell(row=row, column=col)
                novo = ws_novo.cell(row=row, column=col)
                coord = original.coordinate

                if coord in esperado:
                    if _comparavel(novo.value) != _comparavel(esperado[coord]):
                        problemas.append(
                            f"  {planilha.aba}!{coord}: gravado={novo.value!r} "
                            f"esperado={esperado[coord]!r}"
                        )
                    continue

                conferidas += 1
                if original.value != novo.value:
                    problemas.append(
                        f"  {planilha.aba}!{coord}: modelo={original.value!r} "
                        f"virou {novo.value!r}"
                    )

        for nome in wb_orig.sheetnames:
            if nome == planilha.aba:
                continue
            oa, ob = wb_orig[nome], wb_novo[nome]
            for linha_o, linha_n in zip(oa.iter_rows(), ob.iter_rows()):
                for original, novo in zip(linha_o, linha_n):
                    conferidas += 1
                    if original.value != novo.value:
                        problemas.append(
                            f"  {nome}!{original.coordinate}: modelo={original.value!r} "
                            f"virou {novo.value!r}"
                        )

        if problemas:
            raise WorkbookError("planilha final divergiu do modelo:\n" + "\n".join(problemas))
        return conferidas
    finally:
        wb_orig.close()
        wb_novo.close()


def _celulas_escritas(
    reference_date: date,
    fills: list[RowFill],
    planilha: PlanilhaConfig,
) -> dict[str, object]:
    """Mapa coordenada -> valor esperado, para as celulas que o programa escreve."""
    esperado: dict[str, object] = {planilha.celula_data: reference_date}
    for fill in fills:
        esperado[f"{planilha.col_saldo}{fill.row}"] = to_float(fill.balance)
        esperado[f"{planilha.col_pagamento}{fill.row}"] = to_float(fill.total)
        esperado[f"{planilha.col_qtd_sistema}{fill.row}"] = fill.count
        esperado[f"{planilha.col_qtd_banco}{fill.row}"] = fill.bank_count
    return esperado
