"""Quem entra em cada conta — lido da aba «Regras» do MODELO.xlsx.

Estas regras NAO moram no config.yaml de proposito. Elas citam pessoa fisica
(os investidores), e o repositorio e publico: o MODELO.xlsx esta no .gitignore,
o config.yaml tambem, mas o modelo e o unico dos dois que o Gustavo edita no
Excel — que e onde a regra muda. Uma copia em YAML seria uma segunda verdade
para divergir da primeira.

A aba «Regras» tem uma linha por par (conta que recebe, aportador):

    B: CONTA QUE RECEBE   C: APORTADOR   D: PESO   E: TIPO   F: TEM CONTA NO PAINEL?

PESO vazio quer dizer "sem regra fixa": o Gustavo divide o valor na hora, e o
programa nao chuta a divisao — ele diz que ela esta em aberto.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import openpyxl

ZERO = Decimal("0")

#: Nome da aba e primeira linha de dados. Ficam aqui, e nao no config.yaml,
#: porque quem os muda e quem edita o modelo — e o modelo e este arquivo.
ABA = "Regras"
PRIMEIRA_LINHA = 5


@dataclass(frozen=True)
class Regra:
    destino: str
    aportador: str
    peso: Decimal | None
    tipo: str
    do_painel: bool  # o aportador tem conta no painel (sai de uma conta sua)


def ler_regras(modelo: str | Path) -> list[Regra]:
    """Le a aba «Regras». Devolve [] se a aba nao existir — nao e obrigatoria."""
    caminho = Path(modelo)
    if not caminho.is_file():
        return []
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        if ABA not in wb.sheetnames:
            return []
        ws = wb[ABA]
        regras: list[Regra] = []
        for linha in ws.iter_rows(min_row=PRIMEIRA_LINHA, min_col=2, max_col=6,
                                  values_only=True):
            destino, aportador, peso, tipo, do_painel = linha
            if not destino or not aportador:
                continue
            regras.append(
                Regra(
                    destino=str(destino).strip(),
                    aportador=str(aportador).strip(),
                    peso=Decimal(str(peso)) if peso not in (None, "") else None,
                    tipo=str(tipo or "Aporte").strip(),
                    do_painel=str(do_painel or "").strip().lower().startswith("s"),
                )
            )
        return regras
    finally:
        wb.close()


def saidas_estimadas(
    minimos: dict[str, Decimal],
    regras: list[Regra],
) -> tuple[dict[str, Decimal], list[tuple[str, Decimal, tuple[str, ...]]]]:
    """Se cada conta receber o MINIMO, quanto cada aportador teria de mandar?

    Devolve (quanto cada aportador manda, contas cujo rateio esta em aberto).

    Uma conta so entra no primeiro se TODOS os seus aportadores tiverem peso:
    peso pela metade nao e rateio, e um rateio pela metade viraria um numero
    com cara de certo. Sem peso, a conta vai para a segunda lista com o valor
    inteiro e os nomes de quem participa — a divisao e decisao do Gustavo.
    """
    por_destino: dict[str, list[Regra]] = {}
    for regra in regras:
        por_destino.setdefault(regra.destino, []).append(regra)

    saidas: dict[str, Decimal] = {}
    a_dividir: list[tuple[str, Decimal, tuple[str, ...]]] = []

    for destino, valor in minimos.items():
        participantes = por_destino.get(destino)
        if not participantes:
            continue
        pesos = [r.peso for r in participantes]
        if any(p is None for p in pesos):
            a_dividir.append((destino, valor, tuple(r.aportador for r in participantes)))
            continue
        total_peso = sum(pesos, ZERO)
        if total_peso <= ZERO:
            a_dividir.append((destino, valor, tuple(r.aportador for r in participantes)))
            continue
        for regra in participantes:
            parte = (valor * regra.peso / total_peso).quantize(Decimal("0.01"))
            saidas[regra.aportador] = saidas.get(regra.aportador, ZERO) + parte

    return saidas, a_dividir
