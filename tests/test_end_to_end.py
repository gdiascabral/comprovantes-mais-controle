"""Snapshot -> planilha + resumo, sem browser.

Estes testes usam um snapshot SINTETICO (nao sao dados reais do ERP): o objetivo
e provar o encadeamento completo do pipeline offline.
"""

import dataclasses
import shutil
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from conciliacao.models import ErpAccount, ErpPayment, Snapshot
from conciliacao.pipeline import analyze, run_offline
from conciliacao.validate import Nivel, ValidationError, erros
from conciliacao.workbook import output_name

HOJE = date(2026, 7, 30)


@pytest.fixture
def config_tmp(tmp_path, config, modelo_path):
    """Config apontando para um diretorio temporario, com o modelo copiado."""
    shutil.copy2(modelo_path, tmp_path / "MODELO.xlsx")
    return dataclasses.replace(config, raiz=tmp_path)


def snapshot_de(mapping, *, saldos=None, pagamentos=(), agregado=None) -> Snapshot:
    """Monta um snapshot com TODAS as contas vivas do painel presentes no ERP.

    `saldos` mapeia linha -> saldo; as demais recebem R$ 10.000.
    """
    saldos = saldos or {}
    accounts = [
        ErpAccount(
            id=f"uuid-{row.row}",
            name=row.erp_name,
            account_number=row.account_number,
            balance=saldos.get(row.row, Decimal("10000")),
        )
        for row in mapping.live_rows
    ]
    return Snapshot(
        reference_date=HOJE,
        collected_at="2026-07-30T08:00:00",
        accounts=accounts,
        payments=list(pagamentos),
        page_aggregate_open=agregado,
    )


def pagamento(conta, valor, *, favorecido="FORNECEDOR", status="Em aberto", venc=HOJE):
    return ErpPayment(
        due_date=venc,
        status=status,
        amount=Decimal(valor),
        payee=favorecido,
        account_label=conta,
    )


def test_dia_completo_gera_planilha_e_resumo(config_tmp, mapping):
    snap = snapshot_de(
        mapping,
        saldos={8: Decimal("50000"), 10: Decimal("300")},
        pagamentos=[
            pagamento("MORAIS ENGENHARIA - INTER", "45678.90"),
            pagamento("TERRA BELA - SICOOB", "1000.00"),
            pagamento("TERRA BELA - SICOOB", "1.00"),  # excluido pela regra do R$ 1
        ],
    )

    resultado = run_offline(snap, config_tmp, mapping)

    assert resultado.arquivo.name == output_name(HOJE)
    assert resultado.arquivo.is_file()

    ws = openpyxl.load_workbook(resultado.arquivo).worksheets[0]
    assert ws["D8"].value == pytest.approx(50000)
    assert ws["E8"].value == pytest.approx(45678.90)
    assert ws["I8"].value == 1
    assert ws["E10"].value == pytest.approx(1000.0)
    assert ws["I10"].value == 1
    # Formula preservada.
    assert ws["G8"].value == "=D8-E8-F8"

    # Terra Bela tem 300 e deve pagar 1000 -> precisa de 700.
    assert "TERRA BELA" in resultado.resumo.upper()
    assert "700,00" in resultado.resumo
    assert "APORTE" in resultado.resumo.upper()


def test_dia_sem_pagamentos_zera_tudo(config_tmp, mapping):
    resultado = run_offline(snapshot_de(mapping), config_tmp, mapping)

    ws = openpyxl.load_workbook(resultado.arquivo).worksheets[0]
    for row in (8, 10, 26):
        assert ws[f"E{row}"].value == 0
        assert ws[f"I{row}"].value == 0
        assert ws[f"J{row}"].value == 0  # regra do zero

    assert "Nenhuma conta precisa de aporte hoje" in resultado.resumo


def test_rateio_do_buritis_aparece_no_resumo(config_tmp, mapping):
    """Deficit no Buritis - Inter deve explicitar a parte da Julio/Livian."""
    snap = snapshot_de(
        mapping,
        saldos={9: Decimal("0")},
        pagamentos=[pagamento("MORAIS EMPREENDIMENTOS BURITIS - INTER", "30000.00")],
    )
    resultado = run_offline(snap, config_tmp, mapping)

    assert "Julio/Livian" in resultado.resumo
    assert "10.005,00" in resultado.resumo  # metade dos 20.010 da Morais
    assert "dobro" in resultado.resumo


def test_saldo_nao_lido_impede_a_geracao(config_tmp, mapping):
    """Saldo vazio faria o painel pedir aporte do valor inteiro — tem que travar."""
    snap = snapshot_de(mapping, saldos={8: None})
    snap.payments.append(pagamento("MORAIS ENGENHARIA - INTER", "5000.00"))

    with pytest.raises(ValidationError, match="ficou sem saldo"):
        run_offline(snap, config_tmp, mapping)

    assert not (config_tmp.caminho("saida") / output_name(HOJE)).exists()


def test_forcar_gera_apesar_do_erro(config_tmp, mapping):
    snap = snapshot_de(mapping, saldos={8: None})
    snap.payments.append(pagamento("MORAIS ENGENHARIA - INTER", "5000.00"))

    resultado = run_offline(snap, config_tmp, mapping, forcar=True)
    assert resultado.arquivo.is_file()
    assert erros(resultado.issues)


def test_conta_mapeada_ausente_no_erp_e_erro(config_tmp, mapping):
    snap = snapshot_de(mapping)
    snap.accounts.pop()  # remove uma conta que o painel espera

    with pytest.raises(ValidationError, match="nao foi encontrada no ERP"):
        run_offline(snap, config_tmp, mapping)


def test_pagamento_fora_do_painel_vai_para_log(config_tmp, mapping):
    snap = snapshot_de(
        mapping,
        pagamentos=[
            pagamento("PESSOA FISICA - APENAS LANÇAMENTO", "5000.00", favorecido="ALGUEM"),
            pagamento("CONTA NOVA XPTO - SICOOB", "700.00", favorecido="OUTRO"),
        ],
    )
    resultado = run_offline(snap, config_tmp, mapping)

    assert resultado.log_fora_do_painel is not None
    conteudo = resultado.log_fora_do_painel.read_text(encoding="utf-8")
    assert "[ignorada]" in conteudo
    assert "[desconhecida]" in conteudo
    assert "R$ 5.700,00" in conteudo

    # A conta desconhecida tambem precisa aparecer como aviso no resumo.
    avisos_texto = resultado.resumo
    assert "CONTA NOVA XPTO" in avisos_texto


def test_total_do_dia_acima_do_agregado_mensal_e_erro(config_tmp, mapping):
    """Sinal classico de coleta duplicada: o dia passa o 'Em aberto' do mes."""
    snap = snapshot_de(
        mapping,
        pagamentos=[pagamento("MORAIS ENGENHARIA - INTER", "10000.00")],
        agregado=Decimal("5000.00"),
    )
    with pytest.raises(ValidationError, match="passou do 'Em aberto' do mes"):
        run_offline(snap, config_tmp, mapping)


def test_analyze_nao_escreve_arquivo(config_tmp, mapping):
    resultado = analyze(snapshot_de(mapping), config_tmp, mapping)
    assert resultado.arquivo is None
    assert not (config_tmp.caminho("saida")).exists()


def test_conta_nova_no_erp_gera_aviso(config_tmp, mapping):
    snap = snapshot_de(mapping)
    snap.accounts.append(
        ErpAccount(id="uuid-nova", name="EMPREENDIMENTO NOVO - SICOOB", balance=Decimal("999"))
    )
    resultado = analyze(snap, config_tmp, mapping)

    avisos = [i for i in resultado.issues if i.nivel is Nivel.AVISO]
    assert any("EMPREENDIMENTO NOVO" in i.mensagem for i in avisos)
