# -*- coding: utf-8 -*-
"""Regras do relatório de Pagamentos do Dia.

Nenhum dado real: os payloads abaixo têm a FORMA que a API do ERP devolve,
com nomes e números inventados. O repo é público.
"""
import datetime

import pytest

# Import DIRETO, de propósito: com `importorskip` estes testes sumiam em
# silêncio quando `pagamentos_dia` ficava fora do sys.path — e a suíte
# passava sem executá-los. Falhar no import é o comportamento certo.
import relatorio


def anexo(nome, tag=None, ext=".pdf", url=None):
    return {"filename": nome, "tagName": tag, "extension": ext,
            "downloadUrl": url or f"https://exemplo.invalid/{nome}"}


# --------------------------------------------------------------- tipo de pgto
def test_boleto_ganha_de_pix_mesmo_com_o_erp_dizendo_pix():
    """O ERP marca 'Pix' porque o fornecedor tem chave no cadastro, mas o
    título veio com boleto anexado. Pagar por pix duplicaria o pagamento."""
    item = {"tradePayablePaymentMethod": "Pix",
            "paidToBankAccount": "PIX CNPJ: 11.222.333/0001-44"}
    files = [anexo("boleto oc 1234", "Boleto"), anexo("oc 1234", "Recibo")]
    assert relatorio.tipo_de_pagamento(item, files) == "Boleto"


def test_sem_boleto_continua_pix():
    item = {"tradePayablePaymentMethod": "Pix",
            "paidToBankAccount": "PIX CNPJ: 11.222.333/0001-44"}
    assert relatorio.tipo_de_pagamento(item, [anexo("oc 1234", "Recibo")]) == "Pix"


def test_nota_fiscal_nao_conta_como_boleto():
    item = {"tradePayablePaymentMethod": "Pix", "paidToBankAccount": "PIX CPF: 111.222.333-44"}
    assert relatorio.tipo_de_pagamento(item, [anexo("DANFE 999", "Nota Fiscal")]) == "Pix"


def test_extension_vem_com_ponto():
    assert relatorio.eh_pdf({"extension": ".pdf", "filename": "1234-PED-5678"})


def test_nao_chuta_a_nota_como_boleto():
    assert relatorio.escolher_pdf_do_boleto([anexo("DANFE 999", "Nota Fiscal")]) is None


def test_aceita_fatura_sem_a_palavra_boleto():
    """Fatura de concessionária vem com tagName nulo e nome só de número."""
    escolhido = relatorio.escolher_pdf_do_boleto([anexo("000449239501236")])
    assert escolhido and escolhido["filename"] == "000449239501236"


# ------------------------------------------------------------------ descrição
def test_descricao_usa_nf_e_oc():
    item = {"documentNumber": "5909",
            "costCentreDetails": [{"workName": "RPB 24 QD 26A LT 12"}]}
    overview = {"purchaseOrder": {"number": 6510}}
    assert relatorio.monta_descricao(item, [], "", overview) == \
        "RPB 24 QD 26A LT 12 NF 5909 OC 6510"


def test_oc_do_overview_vence_o_nome_do_anexo():
    item = {"documentNumber": "1", "costCentreDetails": [{"workName": "OBRA"}]}
    assert relatorio.achar_oc(item, [anexo("oc 999")], "", {"purchaseOrder": {"number": 6510}}) \
        == "6510"


def test_oc_por_extenso_no_comentario():
    item = {"costCentreDetails": [{"workName": "OBRA"}]}
    assert relatorio.achar_oc(item, [], "Ordem de Compra: 5413") == "5413"


def test_mao_de_obra_vira_contrato_e_medicao():
    item = {"costCentreDetails": [{"workName": "OBRA X"}],
            "description": "Serviço - 4412 - Medição: 7"}
    assert relatorio.monta_descricao(item, []) == "OBRA X C 4412 M 7"


def test_agua_e_luz_usam_a_descricao_e_nao_o_numero_da_fatura():
    item = {"paidTo": "Equatorial", "documentNumber": "2026068284705",
            "description": "UC 451784501210 REF JUL 2026 CASA 1",
            "costCentreDetails": [{"workName": "TB 17 QD 48 LT 38"}]}
    assert relatorio.monta_descricao(item, []) == \
        "TB 17 QD 48 LT 38 UC 451784501210 REF JUL 2026 CASA 1"


def test_descricao_longa_nao_corta_o_que_distingue_as_linhas():
    """Três lançamentos só diferem no 'CASA 1/2/3' lá no fim do texto."""
    item = {"costCentreDetails": [{"workName": "Pos obra"}],
            "description": "SUPRESSAO DE AGUA (casa entregue sem transferencia) - CASA 3"}
    assert "CASA 3" in relatorio.monta_descricao(item, [])


# ----------------------------------------------------------------- chave Pix
@pytest.mark.parametrize("bruto, esperado", [
    ("PIX CNPJ: 11.222.333/0001-44", "11.222.333/0001-44"),
    ("CHAVE PIX : 111.222.333-44", "111.222.333-44"),
    ("PIX CELULAR: (62) 99876-5432", "(62) 99876-5432"),
    ("CHAVE PIX: fulano@exemplo.invalid", "fulano@exemplo.invalid"),
])
def test_extrai_so_a_chave(bruto, esperado):
    assert relatorio.extrair_chave_pix(bruto) == esperado


def test_recado_no_lugar_da_chave_nao_e_chave():
    assert not relatorio.parece_chave_pix("VER COMENTARIO DA SOLICITACAO")
    assert relatorio.parece_chave_pix("111.222.333-44")


def test_copia_e_cola_vem_da_observacao():
    """Pedido de marketplace: o EMV inteiro estava no campo de observação."""
    coment = ("CODIGO PIX: 00020126540014br.gov.bcb.pix0132pix@exemplo.invalid"
              "52040000530398654061.005802BR5910EXEMPLO_06008Sao Paulo6304ABCD")
    achado = relatorio.chave_pix_do_comentario(coment)
    assert achado.startswith("000201") and achado.endswith("6304ABCD")


def test_observacao_sem_pix_nao_inventa_chave():
    assert relatorio.chave_pix_do_comentario("favor pagar até sexta") == ""


# --------------------------------------------------------------- cruzamento
CHAVE_NFE = "52260711222333000144550010000059090001234567"


def test_chave_de_acesso_entrega_numero_e_cnpj():
    d = relatorio.dados_da_chave_nfe(CHAVE_NFE)
    assert d["numero"] == "5909"
    assert d["cnpj"] == "11222333000144"
    assert d["modelo"] == "55"


def test_nf_do_anexo_diferente_do_lancamento_e_divergencia():
    item = {"documentNumber": "999999", "remainingValue": 10.0, "paidTo": "Fornecedor"}
    resumo, divergiu = relatorio.conferir_documento(item, [anexo(CHAVE_NFE)], ["x"])
    assert divergiu and "DIVERGE" in resumo


def test_cnpj_do_pix_diferente_do_emitente_e_divergencia():
    item = {"documentNumber": "5909", "remainingValue": 10.0, "paidTo": "Fornecedor",
            "paidToBankAccount": "PIX CNPJ: 99.888.777/0001-66"}
    resumo, divergiu = relatorio.conferir_documento(item, [anexo(CHAVE_NFE)], ["x"])
    assert divergiu and "CNPJ DIVERGE" in resumo


def test_tudo_batendo_nao_diverge():
    item = {"documentNumber": "5909", "remainingValue": 5020.28, "paidTo": "Tintas Exemplo",
            "paidToBankAccount": "PIX CNPJ: 11.222.333/0001-44"}
    resumo, divergiu = relatorio.conferir_documento(
        item, [anexo(CHAVE_NFE)], ["TINTAS EXEMPLO LTDA  TOTAL 5.020,28"])
    assert not divergiu
    assert "NF 5909 ✓" in resumo and "CNPJ ✓" in resumo
    assert "valor ✓" in resumo and "fornecedor ✓" in resumo


def test_nao_verifiquei_nao_pode_virar_alarme():
    item = {"documentNumber": "", "remainingValue": 77.0, "paidTo": "Fulano"}
    resumo, divergiu = relatorio.conferir_documento(item, [], [])
    assert not divergiu and "?" in resumo


def test_uc_confere_pelo_nome_do_anexo_sem_baixar_pdf():
    item = {"paidTo": "Equatorial", "description": "UC 451784501210 REF JUL 2026",
            "costCentreDetails": [{"workName": "TB 17 QD 48 LT 38"}]}
    resumo, divergiu = relatorio.conferir_documento(item, [anexo("000451784501210")], [])
    assert not divergiu and "UC 451784501210 ✓" in resumo


def test_uc_de_outra_unidade_e_divergencia():
    item = {"paidTo": "Equatorial", "description": "UC 451784501210 REF JUL 2026",
            "costCentreDetails": [{"workName": "TB 17"}]}
    _, divergiu = relatorio.conferir_documento(item, [anexo("000999999999999")], [])
    assert divergiu


def test_endereco_confere_pelo_nome_da_rua():
    """A concessionária escreve o logradouro; o ERP escreve QD/LT."""
    item = {"paidTo": "Equatorial", "description": "UC 449239501236 REF jul 2026",
            "costCentreDetails": [{"workName": "RUA CASSIMIRO MARQUES QD 18 LT 8"}]}
    resumo, _ = relatorio.conferir_documento(
        item, [anexo("000449239501236")], ["CASSIMIRO MARQUES, 100 - CENTRO"])
    assert "endereço ✓" in resumo


# ---------------------------------------------------------------- valor/contas
def test_titulo_quitado_tem_valor_em_sumOfPaidValues():
    assert relatorio.valor_do_item({"remainingValue": 0.0, "sumOfPaidValues": 5020.28}) == 5020.28


def test_titulo_aberto_usa_remainingValue():
    assert relatorio.valor_do_item({"remainingValue": 300.0, "sumOfPaidValues": 0.0}) == 300.0


def test_conta_de_ajuste_fica_de_fora():
    assert not relatorio.conta_entra("PESSOA FISICA - APENAS LANÇAMENTO")
    assert relatorio.conta_entra("TERRA BELA - SICOOB")


@pytest.mark.parametrize("nome", [
    "PESSOA FISICA - APENAS LANÇAMENTO",
    "PESSOA FISICA - APENAS LANCAMENTO",     # sem cedilha
    "pessoa fisica - apenas  lancamento",    # caixa e espaço dobrado
    "CONTA ERRADA",
])
def test_conta_de_ajuste_nao_depende_de_acento(nome):
    """O nome vem do cadastro do ERP, digitado por gente.

    Era a ÚNICA comparação de nome do módulo que ainda casava contra o texto
    cru: escrito sem cedilha, a conta de ajuste escapava da regra e nascia
    MARCADA na tela, com os lançamentos dela entrando como aptos.
    """
    assert not relatorio.conta_entra(nome)


def test_excluir_vence_incluir_e_ignora_acento():
    assert not relatorio.conta_entra("JOÃO VITOR - CONTA PESSOAL",
                                     incluir=["joao"], excluir=["joao vitor"])


def test_abas_com_prefixo_igual_nao_colidem():
    """O Excel corta em 31 caracteres e recusa nome repetido."""
    nomes = relatorio.nomes_de_aba(["MORAIS EMPREENDIMENTOS BURITIS - INTER",
                                    "MORAIS EMPREENDIMENTOS BURITIS - SICOOB"])
    assert len(set(nomes.values())) == 2
    assert all(len(v) <= 31 for v in nomes.values())


def test_periodo_e_conferido_no_cliente():
    """Se a API ignorar o filtro, a linha de fora não pode entrar calada."""
    hoje = datetime.date(2026, 8, 7)
    itens = [{"plannedDate": "2026-08-07"}, {"plannedDate": "2026-08-07T00:00:00Z"},
             {"plannedDate": "2026-09-01"}]
    assert len(relatorio.filtrar_periodo(itens, hoje, hoje, log=lambda *_: None)) == 2


# ------------------------------------------------------------ linha completa
def test_linha_de_boleto_com_aviso_de_pix_no_cadastro():
    item = {"id": "i1", "tradePayableId": "t1", "paidTo": "Fornecedor Exemplo",
            "remainingValue": 1850.0, "tradePayablePaymentMethod": "Pix",
            "paidToBankAccount": "PIX CNPJ: 11.222.333/0001-44",
            "documentNumber": "5909",
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA X"}]}
    anexos = {"t1": [anexo("boleto oc 6510", "Boleto", url="u1")]}
    overviews = {"i1": {"purchaseOrder": {"number": 6510}, "comment": ""}}
    textos = {"u1": "34191.57007 00024.434375 24177.010000 1 99990000185000"}

    reg = relatorio.montar_registros([item], anexos, overviews, textos).contas
    linha = reg["CONTA TESTE"][0]
    assert linha["tipo"] == "Boleto"
    assert linha["dados"].startswith("34191.57007")
    assert linha["descricao"] == "OBRA X NF 5909 OC 6510"
    assert "pagar o boleto" in linha["obs"]


def test_reembolso_sem_chave_cadastrada_vira_atencao():
    item = {"id": "i2", "tradePayableId": "t2", "paidTo": "Concessionaria",
            "remainingValue": 32.28, "documentNumber": "REEMBOLSO",
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA"}]}
    anexos = {"t2": [anexo("PAGAR PARA FULANO", ext=".pdf", url="u2")]}
    reg = relatorio.montar_registros([item], anexos, {}, {"u2": ""}).contas
    linha = reg["CONTA TESTE"][0]
    assert linha["dados"] == ""
    assert linha["status"].startswith("ATEN")
    assert "fulano" in linha["obs"].lower()


def test_reembolso_com_chave_cadastrada():
    item = {"id": "i3", "tradePayableId": "t3", "paidTo": "Concessionaria",
            "remainingValue": 32.28, "documentNumber": "REEMBOLSO",
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA"}]}
    anexos = {"t3": [anexo("PAGAR PARA FULANO", url="u3")]}
    reg = relatorio.montar_registros([item], anexos, {}, {"u3": ""},
                                     pix_reembolso={"FULANO": "Fulano 111.222.333-44"})
    assert reg.contas["CONTA TESTE"][0]["dados"] == "Fulano 111.222.333-44"


def test_excel_sai_com_as_colunas_na_ordem_pedida(tmp_path):
    from openpyxl import load_workbook
    item = {"id": "i4", "tradePayableId": "t4", "paidTo": "Fornecedor",
            "remainingValue": 100.0, "tradePayablePaymentMethod": "Pix",
            "paidToBankAccount": "PIX CPF: 111.222.333-44",
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA"}]}
    reg = relatorio.montar_registros([item], {}, {}, {})
    destino = relatorio.gerar_excel(reg, tmp_path / "saida.xlsx")
    ws = load_workbook(destino).worksheets[0]
    assert [c.value for c in ws[3]][:5] == \
        ["Tipo de Pgto", "Dados do Pgto", "Valor", "Descrição", "Favorecido"]
    assert ws.cell(row=4, column=3).value == 100.0


# ==========================================================================
# Reembolso: quem recebe atravessa até a remessa
# ==========================================================================
# A planilha é o único ponto que tem, ao mesmo tempo, os anexos, o texto lido
# e o lançamento — é ali que se descobre quem recebe. A remessa lê o veredito
# pronto; redescobri-lo seria uma segunda regra sobre a mesma linha.

CPF_DA_PESSOA = "11144477735"          # sintético, DV fechando


def item_com_aviso(url="u-reembolso"):
    return {"id": "i9", "tradePayableId": "t9", "paidTo": "FORNECEDOR SA",
            "remainingValue": 630.0, "tradePayablePaymentMethod": "Pix",
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA"}]}, \
        {"t9": [anexo("PAGAR PARA FULANO DE TAL", url=url)]}


def test_o_reembolso_resolvido_atravessa_com_nome_e_documento():
    item, anexos = item_com_aviso()
    reg = relatorio.montar_registros(
        [item], anexos, {}, {"u-reembolso": ""},
        participantes={"FULANO DE TAL": CPF_DA_PESSOA})
    linha = reg.contas["CONTA TESTE"][0]
    assert linha["reembolso"]
    assert linha["reembolso_nome"] == "FULANO DE TAL"
    assert linha["reembolso_documento"] == CPF_DA_PESSOA
    assert not linha["reembolso_impedimento"]
    # o favorecido do LANÇAMENTO não se perde: a remessa o usa para dizer de
    # que compra o reembolso veio
    assert linha["favorecido"] == "FORNECEDOR SA"


def test_o_reembolso_sem_documento_atravessa_com_o_motivo():
    item, anexos = item_com_aviso()
    reg = relatorio.montar_registros([item], anexos, {}, {"u-reembolso": ""})
    linha = reg.contas["CONTA TESTE"][0]
    assert linha["reembolso"] and not linha["reembolso_documento"]
    assert "FULANO DE TAL" in linha["reembolso_impedimento"]


def test_linha_comum_nao_ganha_campos_de_reembolso():
    item = {"id": "i8", "tradePayableId": "t8", "paidTo": "Fornecedor",
            "remainingValue": 100.0, "tradePayablePaymentMethod": "Pix",
            "paidToBankAccount": "PIX CPF: 111.222.333-44",
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA"}]}
    linha = relatorio.montar_registros([item], {}, {}, {}).contas["CONTA TESTE"][0]
    assert not linha["reembolso"]
    assert linha["reembolso_nome"] == linha["reembolso_documento"] == ""
    assert linha["reembolso_impedimento"] == ""


# ==========================================================================
# Etapa 2: a janela lista o dia inteiro
# ==========================================================================
def _pagamento(ident, conta="CONTA TESTE", pago=False):
    return {"id": ident, "tradePayableId": ident, "paidTo": "FORNECEDOR SA",
            "remainingValue": 10.0, "paid": pago,
            "tradePayableAccount": {"name": conta}}


def test_a_janela_lista_todo_lancamento_das_contas_marcadas():
    """A inversão: o `confirmar_antes.json` deixou de decidir quem aparece.

    Antes, fornecedor fora daquele arquivo não tinha onde ser tirado do dia —
    a não ser desmarcando a conta inteira, junto com tudo o mais que ela tem.
    """
    import pagamentos_frame as frame

    lancamentos = [_pagamento("a"), _pagamento("b"),
                   _pagamento("c", conta="OUTRA CONTA")]
    alvos = frame.alvos_para_confirmar(lancamentos, ["CONTA TESTE"])
    assert [i["id"] for i in alvos] == ["a", "b"]


def test_ja_pago_nao_entra_na_pergunta():
    """Não há o que decidir sobre ele."""
    import pagamentos_frame as frame

    alvos = frame.alvos_para_confirmar(
        [_pagamento("a"), _pagamento("b", pago=True)], ["CONTA TESTE"])
    assert [i["id"] for i in alvos] == ["a"]


def test_conta_nao_marcada_nao_entra_na_pergunta():
    import pagamentos_frame as frame

    assert frame.alvos_para_confirmar([_pagamento("a")], ["OUTRA"]) == []


#: Cadastro de mentira, no formato de `carregar_fornecedores`.
_MARCADA = {"CONCESSIONARIA LUZ": {"so_marcador": True}}


def _marcador(ident, valor=1.00):
    """O lançamento de R$ 1,00 que a concessionária usa para abrir o mês."""
    item = _pagamento(ident)
    item.update(paidTo="CONCESSIONARIA LUZ S/A", remainingValue=valor)
    return item


def test_o_marcador_de_recorrencia_nao_ocupa_a_janela():
    """A queixa de 20/08/2026: três linhas de R$ 1,00, desmarcadas todo dia.

    A etapa 3 já as descartava por valor simbólico — mas ela roda depois, e a
    janela existe para recolher decisão, não para repetir uma já tomada.
    """
    import pagamentos_frame as frame

    alvos = frame.alvos_para_confirmar(
        [_pagamento("a"), _marcador("luz")], ["CONTA TESTE"], _MARCADA)
    assert [i["id"] for i in alvos] == ["a"]


def test_valor_de_verdade_da_mesma_concessionaria_continua_na_janela():
    """A marca é sobre o R$ 1,00, não sobre o nome: a conta de luz aparece."""
    import pagamentos_frame as frame

    alvos = frame.alvos_para_confirmar(
        [_marcador("conta", valor=56.24)], ["CONTA TESTE"], _MARCADA)
    assert [i["id"] for i in alvos] == ["conta"]


def test_sem_cadastro_a_janela_lista_tudo_como_antes():
    """`fornecedores` é opcional — sem ele, nada é filtrado."""
    import pagamentos_frame as frame

    alvos = frame.alvos_para_confirmar([_marcador("luz")], ["CONTA TESTE"])
    assert [i["id"] for i in alvos] == ["luz"]
