"""Guarda-corpo entre o MODELO.xlsx e o que o programa assume dele.

Se alguem editar as formulas do Excel, estes testes falham e apontam o que
precisa ser atualizado no config.yaml / mapping.yaml. Sem isso, o resumo em
texto passaria a mentir em silencio.
"""

import openpyxl
import pytest


@pytest.fixture(scope="module")
def ws(modelo_path):
    return openpyxl.load_workbook(modelo_path).worksheets[0]


def test_uma_aba_e_faixa_esperada(ws, planilha):
    # Duas linhas depois da última conta: o rateio secundário e o total. Sai
    # do config e não de um número fixo — acrescentar conta ao painel é rotina
    # (em 17/08/2026 entraram duas), e um número aqui obrigaria a lembrar
    # deste arquivo toda vez.
    assert ws.max_row == planilha.ultima_linha + 2
    assert ws[f"B{planilha.primeira_linha}"].value == "Morais Engenharia - Inter"
    # A última linha existe e é uma conta — o nome dela muda quando o painel
    # cresce, então o que se cobra é que ela esteja preenchida.
    assert (ws[f"B{planilha.ultima_linha}"].value or "").strip()


def test_labels_do_mapping_batem_com_a_coluna_b(ws, mapping):
    """B e chave do SUMIF da coluna F — divergencia zera aporte sem erro."""
    for row in mapping.rows:
        assert ws[f"B{row.row}"].value == row.label, f"linha {row.row}"


def test_colunas_preenchidas_estao_vazias_no_modelo(ws, planilha):
    for row in planilha.linhas:
        for col in planilha.colunas_escritas:
            assert ws[f"{col}{row}"].value is None, f"{col}{row}"
    assert ws[planilha.celula_data].value is None


def test_todas_as_colunas_de_formula_tem_formula(ws, planilha):
    """Excecao conhecida: F12 e F13 sao o literal 0, e N e vazio fora da 12."""
    for row in planilha.linhas:
        for col in planilha.colunas_formula:
            valor = ws[f"{col}{row}"].value
            if col == "F" and row in planilha.linhas_com_aporte_zero_fixo:
                assert valor == 0, f"F{row} deveria ser o literal 0"
            elif col == "N":
                esperado = row in planilha.aportes_direcionados
                assert bool(valor) == esperado, f"N{row}"
            else:
                assert isinstance(valor, str) and valor.startswith("="), f"{col}{row}"


def test_formula_de_aportes_recebidos_usa_sumif_sobre_b(ws, planilha):
    for row in planilha.linhas:
        if row in planilha.linhas_com_aporte_zero_fixo:
            continue
        # O intervalo vai até a ÚLTIMA linha de conta do config: conta que
        # entra fora do intervalo aparece no painel e não soma aporte nenhum,
        # sem erro na tela.
        primeira, ultima = planilha.primeira_linha, planilha.ultima_linha
        assert ws[f"F{row}"].value == (
            f"=SUMIF($N${primeira}:$N${ultima},B{row},"
            f"$M${primeira}:$M${ultima})")


def test_formula_de_saldo_final(ws, planilha):
    for row in planilha.linhas:
        assert ws[f"G{row}"].value == f"=D{row}-E{row}-F{row}"


def test_apenas_a_linha_com_rateio_tem_fator_no_aporte_minimo(ws, planilha):
    """Fator != 1 so existe nas linhas declaradas em `rateios` no config."""
    for row in planilha.linhas:
        formula = ws[f"M{row}"].value
        regra = planilha.rateio_da_linha(row)
        if regra is None:
            assert formula == (
                f"=IF(E{row}=\"\",0,IF(D{row}>(E{row}+F{row}),\"-\","
                f"(((E{row}+F{row})-D{row}))))"
            ), f"M{row} tem fator inesperado"
        else:
            assert f"*{regra.fator_principal}" in formula, f"M{row}"


def test_rateio_secundario_e_metade_do_principal(ws, planilha):
    """F32 = M9/2: a Julio/Livian entra com metade do que a Morais poe (2:1)."""
    for regra in planilha.rateios:
        assert ws[regra.celula_secundario].value == (
            f"=IF(M{regra.linha}>0,M{regra.linha}/{regra.divisor_secundario},0)"
        )
        # H da linha rateada soma os dois aportes, fechando 100% do deficit.
        assert ws[f"H{regra.linha}"].value == (
            f"=IF(G{regra.linha}<0,G{regra.linha}+M{regra.linha}"
            f"+{regra.celula_secundario},G{regra.linha})"
        )


def test_aportador_secundario_esta_rotulado_na_planilha(ws, planilha):
    for regra in planilha.rateios:
        linha_rotulo = int(regra.celula_secundario[1:])
        assert ws[f"B{linha_rotulo}"].value == regra.aportador_secundario


def test_saldo_pos_aporte_das_linhas_sem_rateio(ws, planilha):
    for row in planilha.linhas:
        if planilha.rateio_da_linha(row) is not None:
            continue
        assert ws[f"H{row}"].value == f"=IF(G{row}<0,G{row}+M{row},G{row})"


def test_aporte_direcionado_aponta_para_o_label_configurado(ws, planilha):
    for origem, destino in planilha.aportes_direcionados.items():
        formula = ws[f"N{origem}"].value
        assert destino in formula, f"N{origem} = {formula!r}"


def test_total_de_pagamentos(ws, planilha):
    assert ws[planilha.celula_total_pagamentos].value == (
        f"=SUM(E{planilha.primeira_linha}:E{planilha.ultima_linha})"
    )


def test_coluna_l_depende_de_j_e_por_isso_nao_serve_ao_resumo(ws, planilha):
    """Documenta a razao de o resumo usar M/G em vez de L.

    L = IF(J="","-",...). Como J fica vazio quando ha pagamento no dia, L sai
    sempre "-" no arquivo recem-gerado.
    """
    for row in planilha.linhas:
        assert ws[f"L{row}"].value.startswith(f'=IF(J{row}=""')
