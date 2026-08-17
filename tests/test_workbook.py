from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from conciliacao.models import RowFill
from conciliacao.workbook import WorkbookError, build, output_name

#: Onde o rateio secundário (Julio/Livian) mora hoje. Vem do config e
#: não escrito à mão: a célula desce toda vez que uma conta entra no
#: painel, e em 17/08/2026 ela desceu de F32 para F34.
def _celula_rateio():
    from conciliacao.config import load_config
    from pathlib import Path as _P
    cfg = load_config(_P(__file__).resolve().parent.parent / 'config.yaml')
    return cfg.planilha.rateios[0].celula_secundario


CELULA_RATEIO = _celula_rateio()


HOJE = date(2026, 7, 30)


@pytest.fixture
def fills(planilha, mapping):
    """Um dia tipico: uma conta com pagamento, uma sem, e as linhas mortas."""
    resultado = []
    for row in planilha.linhas:
        modelo = mapping.by_row(row)
        if modelo is not None and not modelo.exists_in_erp:
            resultado.append(RowFill(row, None, None, None, None))
        elif row == 8:
            resultado.append(RowFill(row, Decimal("1234567.89"), Decimal("45678.90"), 35, None))
        elif row == 10:
            resultado.append(RowFill(row, Decimal("-1500.75"), Decimal("0"), 0, 0))
        else:
            resultado.append(RowFill(row, Decimal("100"), Decimal("0"), 0, 0))
    return resultado


def test_output_name_usa_dia_e_mes(planilha):
    assert output_name(date(2026, 7, 30)) == "30 07 - completa.xlsx"
    assert output_name(date(2026, 12, 5)) == "05 12 - completa.xlsx"


def test_build_escreve_valores_e_preserva_formulas(
    modelo_path, tmp_path, fills, mapping, planilha
):
    destino = tmp_path / output_name(HOJE)
    resultado = build(modelo_path, destino, HOJE, fills, mapping, planilha)

    assert resultado.path.is_file()
    assert resultado.linhas_escritas == len(planilha.linhas)
    # assert_untouched conferiu o resto da planilha celula por celula.
    assert resultado.celulas_conferidas > 200

    ws = openpyxl.load_workbook(destino).worksheets[0]

    # O Excel guarda data como serial; o openpyxl devolve datetime.
    assert ws[planilha.celula_data].value.date() == HOJE
    assert ws[planilha.celula_data].number_format == planilha.formato_data

    assert ws["D8"].value == pytest.approx(1234567.89)
    assert ws["E8"].value == pytest.approx(45678.90)
    assert ws["I8"].value == 35
    assert ws["J8"].value is None  # ha pagamento -> J fica para a conferencia manual

    assert ws["D10"].value == pytest.approx(-1500.75)
    assert (ws["E10"].value, ws["I10"].value, ws["J10"].value) == (0, 0, 0)

    # Formulas intactas, incluindo o encadeamento de aportes e o rateio.
    # O intervalo acompanha a ultima linha de conta do config: o painel
    # cresce, e o que se cobra e que a formula tenha sido PRESERVADA
    # pela escrita, nao que ela seja de uma versao especifica.
    p, u = planilha.primeira_linha, planilha.ultima_linha
    assert ws["F8"].value == f"=SUMIF($N${p}:$N${u},B8,$M${p}:$M${u})"
    assert ws["G8"].value == "=D8-E8-F8"
    assert ws["M9"].value.endswith("*0.667)))")
    assert ws["H9"].value == f"=IF(G9<0,G9+M9+{CELULA_RATEIO},G9)"
    assert ws[CELULA_RATEIO].value == "=IF(M9>0,M9/2,0)"
    assert ws["N12"].value == '=IF(M12>0,"MORAIS ENGENHARIA - INTER","")'
    # O total tambem sai do config: ele desce quando o painel cresce, e
    # em 17/08/2026 foi de E33 para E35.
    assert (ws[planilha.celula_total_pagamentos].value
            == f"=SUM(E{p}:E{u})")
    assert ws["F12"].value == 0


def test_linhas_sem_conta_no_erp_ficam_vazias_no_arquivo(
    modelo_path, tmp_path, fills, mapping, planilha
):
    destino = tmp_path / "morta.xlsx"
    build(modelo_path, destino, HOJE, fills, mapping, planilha)
    ws = openpyxl.load_workbook(destino).worksheets[0]

    for row in (30, 31):
        for col in planilha.colunas_escritas:
            assert ws[f"{col}{row}"].value is None, f"{col}{row}"


def test_formatacao_condicional_e_validacoes_sobrevivem(
    modelo_path, tmp_path, fills, mapping, planilha
):
    """Round-trip do openpyxl nao pode perder o que o modelo tem de visual."""
    destino = tmp_path / "estilos.xlsx"
    build(modelo_path, destino, HOJE, fills, mapping, planilha)

    orig = openpyxl.load_workbook(modelo_path).worksheets[0]
    novo = openpyxl.load_workbook(destino).worksheets[0]

    assert len(novo.conditional_formatting._cf_rules) == len(orig.conditional_formatting._cf_rules)
    assert len(novo.data_validations.dataValidation) == len(orig.data_validations.dataValidation)
    assert novo.freeze_panes == orig.freeze_panes
    assert novo.merged_cells.ranges == orig.merged_cells.ranges
    assert novo["D8"].number_format == orig["D8"].number_format


def test_o_modelo_original_nao_e_modificado(modelo_path, tmp_path, fills, mapping, planilha):
    antes = modelo_path.read_bytes()
    build(modelo_path, tmp_path / "saida.xlsx", HOJE, fills, mapping, planilha)
    assert modelo_path.read_bytes() == antes


def test_reexecucao_no_mesmo_dia_sobrescreve(modelo_path, tmp_path, fills, mapping, planilha):
    destino = tmp_path / output_name(HOJE)
    build(modelo_path, destino, HOJE, fills, mapping, planilha)
    resultado = build(modelo_path, destino, HOJE, fills, mapping, planilha)
    assert resultado.path == destino


def test_label_divergente_do_modelo_falha_alto(modelo_path, tmp_path, fills, mapping, planilha):
    """Label errado zeraria o SUMIF da coluna F sem erro nenhum no Excel."""
    alvo = mapping.by_row(8)
    # Guarda o valor ORIGINAL em vez de restaurar um literal: a fixture e de
    # escopo session e o label vem do mapping real, que muda com o tempo —
    # restaurar o literal errado contaminaria os testes seguintes.
    anterior = alvo.label
    object.__setattr__(alvo, "label", "Nome Trocado")
    try:
        with pytest.raises(WorkbookError, match="diverge da coluna B"):
            build(modelo_path, tmp_path / "x.xlsx", HOJE, fills, mapping, planilha)
    finally:
        object.__setattr__(alvo, "label", anterior)


def test_modelo_inexistente_falha_com_mensagem_clara(tmp_path, fills, mapping, planilha):
    with pytest.raises(WorkbookError, match="modelo nao encontrado"):
        build(tmp_path / "nao-existe.xlsx", tmp_path / "y.xlsx", HOJE, fills, mapping, planilha)
