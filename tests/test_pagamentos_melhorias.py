# -*- coding: utf-8 -*-
"""Regras novas dos Pagamentos do Dia (conferência de 11/08/2026).

Quem confere anotou, linha a linha, o que a planilha trazia de mais e o que
trazia de menos. Cada teste aqui guarda uma dessas anotações.

Nenhum dado real: os payloads têm a FORMA que a API do ERP devolve, com nomes
e números inventados. O repo é público.
"""
import pytest

# Import DIRETO, de propósito: com `importorskip` estes testes sumiriam em
# silêncio se `pagamentos_dia` saísse do sys.path, e a suíte passaria sem
# executá-los. Falhar no import é o comportamento certo.
import ocr_boleto
import regras_pagamento as regras
import relatorio

#: Linhas digitáveis com dígitos verificadores que fecham DE VERDADE — é o que
#: estes testes precisam provar. Uma linha digitável não diz quem pagou.
LINHA_BANCARIA = "34191.57007 00024.924375 24177.010006 9 15340000115000"
VALOR_BANCARIA = 1150.00
LINHA_ARRECADACAO = "86860000026-5 70860161209-4 22026081001-8 61001177300-1"
VALOR_ARRECADACAO = 2670.86


def anexo(nome, tag=None, ext=".pdf", url=None):
    return {"filename": nome, "tagName": tag, "extension": ext,
            "downloadUrl": url or f"https://exemplo.invalid/{nome}"}


def linhas(resultado):
    """As linhas da única conta do resultado."""
    return next(iter(resultado.contas.values()))


def lancamento(**extra):
    item = {"id": "x1", "tradePayableId": "x1", "paidTo": "Fornecedor Exemplo",
            "remainingValue": 100.0,
            "tradePayableAccount": {"name": "CONTA TESTE"},
            "costCentreDetails": [{"workName": "OBRA X"}]}
    item.update(extra)
    return item


# ==========================================================================
# 1. "não precisa trazer o endereço duas vezes na descrição"
# ==========================================================================
def test_rateio_no_mesmo_imovel_nao_repete_o_endereco():
    """Lançamento rateado repete o MESMO centro de custo em cada parcela, e a
    descrição saía "TB 21 QD 51 LT 38 | TB 21 QD 51 LT 38 NF 61106"."""
    item = {"documentNumber": "61106",
            "costCentreDetails": [{"workName": "TB 21 QD 51 LT 38"},
                                  {"workName": "TB 21 QD 51 LT 38"}]}
    assert relatorio.monta_descricao(item, []) == "TB 21 QD 51 LT 38 NF 61106"


def test_rateio_entre_imoveis_diferentes_continua_mostrando_os_dois():
    """Dois imóveis é informação, não repetição."""
    item = {"documentNumber": "77",
            "costCentreDetails": [{"workName": "QD 51 LT 38"},
                                  {"workName": "QD 51 LT 39"}]}
    assert relatorio.monta_descricao(item, []) == "QD 51 LT 38 | QD 51 LT 39 NF 77"


# ==========================================================================
# 3. "quando o valor tiver 1,00 não precisa trazer" / "0,01 não precisa"
# ==========================================================================
@pytest.mark.parametrize("valor", [1.00, 0.01])
def test_valor_simbolico_nao_entra(valor):
    assert regras.valor_simbolico(valor)
    assert regras.motivo_omissao(valor, "Fulano", "BOLETO", "linha", {}) == \
        regras.MOTIVO_SIMBOLICO


@pytest.mark.parametrize("valor", [0.02, 0.99, 1.01, 2.00, 3.00, 1150.0])
def test_valor_de_verdade_entra(valor):
    """A lista é EXATA, não um piso: taxa de R$ 2,00 é pagamento de verdade."""
    assert not regras.valor_simbolico(valor)
    assert regras.motivo_omissao(valor, "Fulano", "BOLETO", "linha", {}) == ""


def test_linha_de_um_real_vai_para_os_omitidos_com_o_motivo():
    item = lancamento(remainingValue=1.00, tradePayablePaymentMethod="Pix",
                      paidToBankAccount="PIX CNPJ: 11.222.333/0001-44")
    res = relatorio.montar_registros([item], {}, {}, {})
    assert res.contas == {}
    assert [o["motivo"] for o in res.omitidos] == [regras.MOTIVO_SIMBOLICO]
    assert res.omitidos[0]["valor"] == 1.00


def test_ja_pago_escapa_das_regras_de_omissao():
    """Linha já paga é informação, não pagamento: "sem forma de pagar" ali é o
    normal, e omiti-la esconderia justo o que a pessoa pediu para ver."""
    item = lancamento(paid=True, dateOfPayment="2026-08-11", remainingValue=1.00)
    res = relatorio.montar_registros([item], {}, {}, {})
    assert res.omitidos == []
    assert linhas(res)[0]["status"].startswith("JÁ PAGO")


# ==========================================================================
# 2 e 4. sem boleto anexado: paga pelo Pix do cadastro, ou não entra
# ==========================================================================
def sem_boleto(**extra):
    base = {"paidTo": "Calhas Modelo", "remainingValue": 1891.0,
            "tradePayablePaymentMethod": "Boleto",
            "paidToBankAccount": "PIX CNPJ: 22.333.444/0001-55",
            "costCentreDetails": [{"workName": "CONDOMINIO RESERVA"}]}
    base.update(extra)
    return lancamento(**base)


def test_sem_boleto_mas_com_nf_paga_pela_chave_do_cadastro():
    """Sem boleto anexado, "boleto ganha de Pix" não tem premissa: não existe
    boleto para ganhar. Com NF, a compra está documentada e o Pix serve."""
    res = relatorio.montar_registros([sem_boleto(documentNumber="113")], {}, {}, {})
    linha = linhas(res)[0]
    assert linha["tipo"] == "Pix"
    assert linha["dados"] == "22.333.444/0001-55"
    assert "chave Pix do cadastro" in linha["obs"]
    assert "pagar o boleto" not in linha["obs"]


def test_sem_boleto_com_oc_tambem_paga_pela_chave():
    res = relatorio.montar_registros(
        [sem_boleto()], {}, {"x1": {"purchaseOrder": {"number": 6516}}}, {})
    assert linhas(res)[0]["dados"] == "22.333.444/0001-55"


def test_sem_boleto_e_sem_nf_nem_oc_nao_entra():
    """Locação sem nota: nada documenta a compra, e a linha vira ruído."""
    item = sem_boleto(paidTo="Containers Modelo",
                      description="Locacao de conteiner 08/08 a 08/09")
    res = relatorio.montar_registros([item], {}, {}, {})
    assert res.contas == {}
    assert res.omitidos[0]["motivo"] == regras.MOTIVO_SEM_PAGAR


def test_boleto_anexado_continua_ganhando_do_pix():
    """A trava antiga não pode ter afrouxado: com boleto na mão paga-se o
    boleto, senão o mesmo título é pago duas vezes."""
    anexos = {"x1": [anexo("boleto oc 6516", "Boleto", url="ub")]}
    res = relatorio.montar_registros([sem_boleto(documentNumber="113")], anexos,
                                     {}, {"ub": LINHA_BANCARIA})
    linha = linhas(res)[0]
    assert linha["tipo"] == "Boleto"
    assert linha["dados"].startswith("34191.57007")
    assert "pagar o boleto" in linha["obs"]


def test_sem_chave_e_sem_aviso_nao_entra():
    item = lancamento(tradePayablePaymentMethod="Cartão de crédito",
                      documentNumber="740")
    res = relatorio.montar_registros([item], {}, {}, {})
    assert res.contas == {}
    assert res.omitidos[0]["motivo"] == regras.MOTIVO_SEM_PAGAR


def test_reembolso_sem_chave_mas_com_aviso_continua_entrando():
    """"não tem chave para pagar, não tem aviso para pagar" são as DUAS
    coisas. Com aviso anexado, alguém escreveu aquilo hoje para pagar."""
    item = lancamento(documentNumber="REEMBOLSO")
    anexos = {"x1": [anexo("PAGAR PARA FULANO", url="ur")]}
    res = relatorio.montar_registros([item], anexos, {}, {"ur": ""})
    assert res.omitidos == []
    assert linhas(res)[0]["status"].startswith("ATEN")


# ==========================================================================
# 9. "esse fornecedor sempre será reembolso" (são dois, e ficam no cadastro)
# ==========================================================================
REGRAS_FORN = {"VIDRACARIA MODELO": {"so_com_reembolso": True},
               "MADEIREIRA MODELO": {"so_com_reembolso": True}}


def test_fornecedor_de_reembolso_sem_aviso_nao_entra():
    item = lancamento(paidTo="VIDRACARIA MODELO LTDA", remainingValue=5020.0,
                      tradePayablePaymentMethod="Pix",
                      paidToBankAccount="PIX CNPJ: 11.222.333/0001-44")
    res = relatorio.montar_registros([item], {}, {}, {}, regras_fornecedor=REGRAS_FORN)
    assert res.contas == {}
    assert res.omitidos[0]["motivo"] == regras.MOTIVO_REEMBOLSO


def test_fornecedor_de_reembolso_com_aviso_entra():
    item = lancamento(paidTo="MADEIREIRA MODELO", remainingValue=800.0)
    anexos = {"x1": [anexo("PAGAR PARA FULANO", url="uv")]}
    res = relatorio.montar_registros([item], anexos, {}, {"uv": ""},
                                     regras_fornecedor=REGRAS_FORN,
                                     pix_reembolso={"FULANO": "111.222.333-44"})
    assert linhas(res)[0]["dados"] == "111.222.333-44"


def test_regra_do_fornecedor_casa_por_pedaco_e_ignora_acento():
    """O nome vem do cadastro do ERP e é digitado por gente."""
    assert regras.regra_do_fornecedor("MADEIREIRA MODELO LTDA ME", REGRAS_FORN)
    assert regras.regra_do_fornecedor("Vidracaria  Modelo Comercio", REGRAS_FORN)
    assert not regras.regra_do_fornecedor("Vidraçaria Outra", REGRAS_FORN)


def test_regra_mais_especifica_vence():
    r = {"SERVICOS MODELO": {"conferir_endereco": True},
         "SERVICOS MODELO PINTURAS": {"so_com_reembolso": True}}
    assert regras.regra_do_fornecedor("SERVIÇOS MODELO PINTURAS LTDA", r) == \
        {"so_com_reembolso": True}


def test_cadastro_ausente_nao_derruba_o_relatorio(tmp_path):
    """Sem os arquivos o app roda igual, só sem as regras."""
    assert regras.carregar_fornecedores(tmp_path) == {}
    assert regras.carregar_confirmar(tmp_path) == []


def test_cadastro_ilegivel_tambem_nao_derruba(tmp_path):
    (tmp_path / regras.ARQ_FORNECEDORES).write_text("{ isto não é json",
                                                    encoding="utf-8")
    assert regras.carregar_fornecedores(tmp_path) == {}


def test_cadastro_e_lido_do_arquivo(tmp_path):
    (tmp_path / regras.ARQ_FORNECEDORES).write_text(
        '{"Vidracaria Modelo": {"so_com_reembolso": true}}', encoding="utf-8")
    (tmp_path / regras.ARQ_CONFIRMAR).write_text(
        '{"nomes": ["Fulano de Tal", "Sicrano de Tal"]}', encoding="utf-8")
    assert regras.carregar_fornecedores(tmp_path)["VIDRACARIA MODELO"]["so_com_reembolso"]
    assert regras.exige_confirmacao("FULANO DE TAL SOCIO",
                                    regras.carregar_confirmar(tmp_path))
    assert not regras.exige_confirmacao("Fornecedor Exemplo",
                                        regras.carregar_confirmar(tmp_path))


# ==========================================================================
# 6. "sempre que for pagamento para os sócios, me perguntar"
# ==========================================================================
def socio(**extra):
    return lancamento(paidTo="FULANO DE TAL SOCIO", id="jv",
                      tradePayableId="jv", remainingValue=50000.0,
                      tradePayablePaymentMethod="Pix",
                      paidToBankAccount="PIX CPF: 111.222.333-44", **extra)


def test_nao_confirmado_sai_com_motivo_e_nao_some():
    res = relatorio.montar_registros([socio()], {}, {}, {},
                                     ids_nao_confirmados={"jv"})
    assert res.contas == {}
    assert res.omitidos[0]["motivo"] == regras.MOTIVO_NAO_CONFIRMADO
    assert res.omitidos[0]["valor"] == 50000.0


def test_confirmado_entra_normalmente():
    res = relatorio.montar_registros([socio()], {}, {}, {}, ids_nao_confirmados=set())
    assert linhas(res)[0]["dados"] == "111.222.333-44"


# ==========================================================================
# 10a. "todo boleto ela coloca no Número do documento o número da OC"
# ==========================================================================
def test_oc_escrita_no_numero_do_documento_nao_vira_nf():
    """Saía "NF 5928 OC 5928": o mesmo número anunciado como duas coisas."""
    item = {"documentNumber": "OC5928", "costCentreDetails": [{"workName": "TB 18"}]}
    assert relatorio.monta_descricao(item, [], "", {"purchaseOrder": {"number": 5928}}) \
        == "TB 18 OC 5928"


def test_documento_igual_a_oc_tambem_nao_vira_nf():
    item = {"documentNumber": "5928", "costCentreDetails": [{"workName": "TB 18"}]}
    assert relatorio.monta_descricao(item, [], "", {"purchaseOrder": {"number": 5928}}) \
        == "TB 18 OC 5928"


def test_nota_de_verdade_continua_sendo_nf():
    item = {"documentNumber": "124613", "costCentreDetails": [{"workName": "QD 01 LT 16"}]}
    assert relatorio.monta_descricao(item, [], "", {"purchaseOrder": {"number": 5710}}) \
        == "QD 01 LT 16 NF 124613 OC 5710"


def test_oc_do_documento_diferente_da_ordem_e_divergencia():
    """"sempre confirmar o XXXX" — quando não bate, alguém tem de olhar."""
    item = lancamento(documentNumber="OC5928", paidTo="Servicos Modelo",
                      remainingValue=1950.0)
    resumo, divergiu = relatorio.conferir_documento(
        item, [], ["servicos modelo"], {"purchaseOrder": {"number": 6100}})
    assert divergiu and "OC DIVERGE" in resumo


def test_oc_do_documento_batendo_confere():
    item = lancamento(documentNumber="OC5928", paidTo="Servicos Modelo",
                      remainingValue=1950.0)
    resumo, divergiu = relatorio.conferir_documento(
        item, [], ["servicos modelo OBRA X"], {"purchaseOrder": {"number": 5928}})
    assert not divergiu and "OC 5928 ✓" in resumo


# ==========================================================================
# 10b. "nesse anexo tem o endereço. Sempre confirmar se o endereço bate"
# ==========================================================================
def test_endereco_do_pedido_confere_com_o_centro_de_custo():
    item = lancamento(documentNumber="1", paidTo="Servicos Modelo", remainingValue=1950.0,
                      costCentreDetails=[{"workName": "TB 18 QD 49 LT 38"}])
    resumo, divergiu = relatorio.conferir_documento(
        item, [], ["PEDIDO 5928 - OBRA TB 18 QD 49 LT 38 - SERVICOS MODELO"])
    assert "endereço ✓" in resumo and not divergiu


def test_endereco_que_nao_aparece_informa_sem_alarmar():
    """Nunca vira laranja: o anexo escreve "QUADRA 49" onde o ERP põe "QD 49",
    e alarme falso ensina o usuário a ignorar alarme."""
    item = lancamento(documentNumber="1", paidTo="Servicos Modelo", remainingValue=1950.0,
                      costCentreDetails=[{"workName": "TB 18 QD 49 LT 38"}])
    resumo, divergiu = relatorio.conferir_documento(item, [], ["PEDIDO 5928 SERVICOS MODELO"])
    assert "endereço ?" in resumo and not divergiu


def test_endereco_pela_metade_diz_o_que_faltou():
    item = lancamento(documentNumber="1", paidTo="Servicos Modelo", remainingValue=1950.0,
                      costCentreDetails=[{"workName": "TB 18 QD 49 LT 38"}])
    resumo, _ = relatorio.conferir_documento(item, [], ["PEDIDO OBRA QD 49 SERVICOS MODELO"])
    assert "LT 38 não aparece" in resumo


def test_lancamento_sem_quadra_e_lote_nao_ganha_coluna_de_endereco():
    item = lancamento(documentNumber="1", paidTo="INSTITUTO MODELO", remainingValue=120.0,
                      costCentreDetails=[{"workName": "ESCRITORIO"}])
    resumo, _ = relatorio.conferir_documento(item, [], ["INSTITUTO MODELO nota 995928"])
    assert "endereço" not in resumo


# ==========================================================================
# 5. "fazer ocr e ler boleto"
# ==========================================================================
@pytest.mark.parametrize("linha, valor", [(LINHA_BANCARIA, VALOR_BANCARIA),
                                          (LINHA_ARRECADACAO, VALOR_ARRECADACAO)])
def test_linha_digitavel_real_passa_nas_duas_provas(linha, valor):
    assert ocr_boleto.valida(linha)
    assert ocr_boleto.confere_valor(linha, valor)
    assert ocr_boleto.linha_confiavel(linha, valor)


@pytest.mark.parametrize("posicao", [0, 5, 20, 32, 46])
def test_um_digito_trocado_reprova(posicao):
    """É para isto que o dígito verificador existe: um caractere lido errado
    paga a conta de outra pessoa, sem erro na tela e sem volta."""
    d = list(ocr_boleto.digitos(LINHA_BANCARIA))
    d[posicao] = "0" if d[posicao] != "0" else "9"
    assert not ocr_boleto.linha_confiavel("".join(d), VALOR_BANCARIA)


@pytest.mark.parametrize("posicao", [1, 12, 25, 47])
def test_um_digito_trocado_reprova_tambem_na_arrecadacao(posicao):
    d = list(ocr_boleto.digitos(LINHA_ARRECADACAO))
    d[posicao] = "0" if d[posicao] != "0" else "9"
    assert not ocr_boleto.linha_confiavel("".join(d), VALOR_ARRECADACAO)


def test_linha_certa_com_valor_de_outro_lancamento_reprova():
    """Segunda prova: a linha pode estar íntegra e ser de outro título."""
    assert ocr_boleto.valida(LINHA_BANCARIA)
    assert not ocr_boleto.linha_confiavel(LINHA_BANCARIA, 999.99)


def test_acha_a_linha_no_meio_do_texto_do_ocr():
    texto = ("BANCO EXEMPLO S.A.\n"
             f"{LINHA_BANCARIA}\n"
             "Vencimento 11/08/2026   Valor do documento 1.150,00\n")
    assert ocr_boleto.achar_linha_digitavel(texto, VALOR_BANCARIA) == \
        ocr_boleto.formatar(LINHA_BANCARIA)


def test_ocr_trocando_letra_por_digito_ainda_e_recuperado():
    """O OCR lê 'O' onde há '0' e 'S' onde há '5'; o mapa de confusões desfaz
    o engano, e quem julga o chute é o dígito verificador."""
    sujo = ocr_boleto.digitos(LINHA_BANCARIA).replace("0", "O").replace("5", "S")
    assert ocr_boleto.achar_linha_digitavel(f"boleto\n{sujo}\n", VALOR_BANCARIA) == \
        ocr_boleto.formatar(LINHA_BANCARIA)


def test_ocr_ilegivel_nao_inventa_linha():
    assert ocr_boleto.achar_linha_digitavel("3419 xxxx 0002 ????  1534", 1150.0) == ""
    assert ocr_boleto.achar_linha_digitavel("", 1150.0) == ""


def test_boleto_em_imagem_vira_linha_digitavel_quando_o_ocr_fecha():
    item = lancamento(paidTo="CONSULTORIA MODELO", remainingValue=VALOR_BANCARIA,
                      documentNumber="537")
    anexos = {"x1": [anexo("boleto consultoria", "Boleto", url="uo")]}
    res = relatorio.montar_registros([item], anexos, {},
                                     {"uo": f"CONSULTORIA MODELO\n{LINHA_BANCARIA}\n"},
                                     urls_ocr={"uo"})
    linha = linhas(res)[0]
    assert linha["dados"] == ocr_boleto.formatar(LINHA_BANCARIA)
    assert "OCR" in linha["obs"]
    assert not linha["status"].startswith("ATEN")


def test_ocr_que_nao_fecha_continua_pedindo_preenchimento_manual():
    """Recusar leitura duvidosa é a única falha aceitável aqui."""
    item = lancamento(paidTo="CONSULTORIA MODELO", remainingValue=11262.0,
                      documentNumber="537")
    anexos = {"x1": [anexo("boleto consultoria", "Boleto", url="uo")]}
    res = relatorio.montar_registros([item], anexos, {}, {"uo": "borrado 1234 5678"},
                                     urls_ocr={"uo"})
    linha = linhas(res)[0]
    assert linha["dados"] == ""
    assert "preencher manual" in linha["obs"]
    assert linha["status"].startswith("ATEN")


def test_texto_de_pdf_de_verdade_nao_passa_pela_trava_do_ocr():
    """PDF com camada de texto continua no caminho antigo: a trava do valor
    vale para OCR, e um boleto pago com juros não pode ser recusado."""
    item = lancamento(paidTo="Fornecedor", remainingValue=999.99,
                      documentNumber="537")
    anexos = {"x1": [anexo("boleto", "Boleto", url="up")]}
    res = relatorio.montar_registros([item], anexos, {}, {"up": LINHA_BANCARIA})
    assert linhas(res)[0]["dados"] == LINHA_BANCARIA


# ==========================================================================
# 8. "tem um arquivo falando 'pagar para' e abaixo tem um numero"
# ==========================================================================
AVISO = ("AUTORIZACAO DE REEMBOLSO\n"
         "PAGAR PARA: FULANO DE TAL\n"
         "111.222.333-44\n"
         "Valor 1.101,65\n")


def reembolso(**extra):
    return lancamento(paidTo="Center Pisos", remainingValue=1101.65,
                      documentNumber="REEMBOLSO",
                      costCentreDetails=[{"workName": "QD 18 LT 8"}], **extra)


def test_chave_lida_do_proprio_aviso_dispensa_cadastro():
    """Quem escreve o aviso já escreve a chave; exigir cadastro prévio fazia
    todo reembolso de gente nova cair como "chave não cadastrada"."""
    anexos = {"x1": [anexo("PAGAR PARA FULANO", url="ur")]}
    res = relatorio.montar_registros([reembolso()], anexos, {}, {"ur": AVISO})
    linha = linhas(res)[0]
    assert linha["dados"] == "111.222.333-44"
    assert "próprio aviso" in linha["obs"]


def test_chave_do_aviso_diferente_da_cadastrada_vira_atencao():
    anexos = {"x1": [anexo("PAGAR PARA FULANO", url="ur")]}
    res = relatorio.montar_registros([reembolso()], anexos, {}, {"ur": AVISO},
                                     pix_reembolso={"FULANO": "999.888.777-66"})
    linha = linhas(res)[0]
    assert linha["status"].startswith("ATEN")
    assert linha["dados"] == "111.222.333-44"      # o aviso é o documento do dia
    assert "difere da cadastrada" in linha["obs"]


def test_mesma_chave_escrita_de_dois_jeitos_nao_e_divergencia():
    anexos = {"x1": [anexo("PAGAR PARA FULANO", url="ur")]}
    res = relatorio.montar_registros([reembolso()], anexos, {}, {"ur": AVISO},
                                     pix_reembolso={"FULANO": "Fulano 11122233344"})
    assert not linhas(res)[0]["status"].startswith("ATEN")


def test_numero_longe_do_pagar_para_nao_e_confundido_com_a_chave():
    """O documento traz também o CNPJ da loja e o valor; a busca fica na
    janela logo depois do "PAGAR PARA", que é onde a chave é escrita."""
    texto = "NOTA FISCAL 11.222.333/0001-44 VALOR 1.101,65\n" * 3
    files = [anexo("PAGAR PARA FULANO", url="ux")]
    assert relatorio.chave_pix_do_aviso(files, {"ux": texto}) == ""


def test_aviso_sem_texto_cai_no_cadastro_como_antes():
    anexos = {"x1": [anexo("PAGAR PARA FULANO", url="ur")]}
    res = relatorio.montar_registros([reembolso()], anexos, {}, {"ur": ""},
                                     pix_reembolso={"FULANO": "999.888.777-66"})
    assert linhas(res)[0]["dados"] == "999.888.777-66"


# ==========================================================================
# A aba "NÃO ENTRARAM": omitir não é apagar
# ==========================================================================
def test_omitidos_viram_aba_propria_com_o_motivo(tmp_path):
    from openpyxl import load_workbook
    bom = lancamento(tradePayablePaymentMethod="Pix",
                     paidToBankAccount="PIX CPF: 111.222.333-44")
    simbolico = lancamento(id="x2", tradePayableId="x2", remainingValue=1.00,
                           paidTo="Equatorial", tradePayablePaymentMethod="Pix",
                           paidToBankAccount="PIX CPF: 111.222.333-44")
    res = relatorio.montar_registros([bom, simbolico], {}, {}, {})
    ws = load_workbook(relatorio.gerar_excel(res, tmp_path / "s.xlsx")
                       )[relatorio.ABA_OMITIDOS]
    assert [c.value for c in ws[3]] == relatorio.HEADERS_OMITIDOS
    assert ws.cell(row=4, column=1).value == "CONTA TESTE"
    assert ws.cell(row=4, column=3).value == 1.00
    assert ws.cell(row=4, column=6).value == regras.MOTIVO_SIMBOLICO


def test_omitido_nao_soma_no_total_da_conta(tmp_path):
    from openpyxl import load_workbook
    bom = lancamento(remainingValue=100.0, tradePayablePaymentMethod="Pix",
                     paidToBankAccount="PIX CPF: 111.222.333-44")
    simbolico = lancamento(id="x2", tradePayableId="x2", remainingValue=1.00,
                           tradePayablePaymentMethod="Pix",
                           paidToBankAccount="PIX CPF: 111.222.333-44")
    res = relatorio.montar_registros([bom, simbolico], {}, {}, {})
    wb = load_workbook(relatorio.gerar_excel(res, tmp_path / "s.xlsx"))
    ws = wb.worksheets[0]
    total = [c.value for c in ws[ws.max_row]]
    assert total[0] == "TOTAL" and total[relatorio.COL_VALOR - 1] == 100.0


def test_sem_omitidos_nao_cria_a_aba(tmp_path):
    from openpyxl import load_workbook
    item = lancamento(tradePayablePaymentMethod="Pix",
                      paidToBankAccount="PIX CPF: 111.222.333-44")
    res = relatorio.montar_registros([item], {}, {}, {})
    wb = load_workbook(relatorio.gerar_excel(res, tmp_path / "s.xlsx"))
    assert relatorio.ABA_OMITIDOS not in wb.sheetnames


def test_conta_nao_rouba_o_nome_da_aba_de_omitidos():
    """Nome de aba repetido derruba a gravação do arquivo inteiro."""
    nomes = relatorio.nomes_de_aba([relatorio.ABA_OMITIDOS, "TERRA BELA"])
    assert nomes[relatorio.ABA_OMITIDOS] != relatorio.ABA_OMITIDOS
    assert len(set(nomes.values())) == 2
